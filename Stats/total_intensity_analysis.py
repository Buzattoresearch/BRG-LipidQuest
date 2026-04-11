from __future__ import annotations

from itertools import combinations
from pathlib import Path
from typing import Dict, List, Optional

import matplotlib as mpl
import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
from matplotlib.colors import to_rgba
from matplotlib.lines import Line2D
from openpyxl import Workbook
from openpyxl.styles import Font
from scipy.stats import ttest_ind

from Stats.figure_style import build_group_palette as _shared_build_group_palette, get_figure_style
from Stats.utils import load_dataset, prepare_output_dir

mpl.rcParams["font.family"] = "Arial"
mpl.rcParams["font.size"] = 18
mpl.rcParams["axes.spines.top"] = False
mpl.rcParams["axes.spines.right"] = False
plt.ioff()

AXIS_GREY = "#7A7A7A"
TICK_GREY = "#565656"
TOP_FADE_COLOR = "#D8D8D8"


def _order_groups(present: List[str], group_order: Optional[List[str]]) -> List[str]:
    present = [str(g) for g in present]
    if not group_order:
        return present
    gui = [g for g in group_order if g in present]
    rest = [g for g in present if g not in gui]
    return gui + rest


def _build_palette(groups_like, group_colors=None, group_order=None):
    return _shared_build_group_palette(groups_like, group_colors=group_colors, group_order=group_order)


def _is_semiquant_dataset(dataset_label: Optional[str], file_path: Optional[str] = None) -> bool:
    dataset_text = str(dataset_label or "").strip().lower()
    file_text = str(file_path or "").strip().lower()
    return "annotated semi-quant" in dataset_text or "semi_quant" in file_text or "semi-quant" in file_text


def _total_intensity_label(dataset_label: Optional[str], file_path: Optional[str] = None) -> str:
    if _is_semiquant_dataset(dataset_label, file_path):
        return "Total summed IS-corrected peak intensities\nfor features annotated as lipids"
    return "Total summed normalized peak intensities\nfor features annotated as lipids"


def _welch_pvalue(x1, x2) -> float:
    x1 = pd.Series(x1).replace([np.inf, -np.inf], np.nan).dropna().astype(float)
    x2 = pd.Series(x2).replace([np.inf, -np.inf], np.nan).dropna().astype(float)
    if len(x1) < 2 or len(x2) < 2:
        return np.nan
    try:
        _, pval = ttest_ind(x1, x2, equal_var=False, nan_policy="omit")
        return float(pval)
    except Exception:
        return np.nan


def _wrap_group_labels(labels: List[str]) -> List[str]:
    wrapped = []
    for label in labels:
        parts = str(label).strip().split()
        if len(parts) <= 1:
            wrapped.append(str(label).strip())
        else:
            wrapped.append(parts[0] + "\n" + " ".join(parts[1:]))
    return wrapped


def _add_y_axis_fade(ax: plt.Axes, ymax: float, x_axis_start: float) -> None:
    if not np.isfinite(ymax) or ymax <= 0:
        return
    y0 = 0.0
    y1 = ymax
    fade_start = ymax * 0.72
    nseg = 80
    ys = np.linspace(y0, y1, nseg + 1)
    for i in range(nseg):
        ya = ys[i]
        yb = ys[i + 1]
        if ya < fade_start:
            alpha = 1.0
        else:
            frac = (ya - fade_start) / max(ymax - fade_start, 1e-12)
            alpha = max(0.0, 1.0 - frac)
        ax.add_line(Line2D(
            [x_axis_start, x_axis_start],
            [ya, yb],
            color=AXIS_GREY,
            linewidth=1.2,
            alpha=alpha,
            zorder=4,
            solid_capstyle="butt",
        ))


def _style_axes(ax: plt.Axes, style: dict) -> None:
    ax.spines["top"].set_visible(False)
    ax.spines["right"].set_visible(False)
    ax.spines["left"].set_visible(False)
    ax.spines["bottom"].set_visible(True)
    ax.spines["bottom"].set_linewidth(style["line_width"])
    ax.spines["bottom"].set_color(AXIS_GREY)
    ax.tick_params(axis="both", colors=TICK_GREY, length=0)
    ax.yaxis.label.set_color(TICK_GREY)
    ax.title.set_color(TICK_GREY)


def _presentation_style(style: dict) -> dict:
    updated = dict(style)
    updated["base_font_size"] = max(int(updated.get("base_font_size", 16)), 18)
    updated["label_size"] = max(int(updated.get("label_size", 16)), 18)
    updated["tick_size"] = max(int(updated.get("tick_size", 16)), 18)
    updated["legend_size"] = max(int(updated.get("legend_size", 16)), 18)
    updated["title_size"] = max(int(updated.get("title_size", 18)), 20)
    updated["line_width"] = max(float(updated.get("line_width", 1.0)), 1.1)
    return updated


def _legend_handles_from_groups(groups: List[str], palette: Dict[str, str]) -> List[Line2D]:
    return [
        Line2D(
            [0],
            [0],
            marker="s",
            linestyle="none",
            markersize=12,
            markerfacecolor=palette[group],
            markeredgecolor="none",
            label=str(group),
        )
        for group in groups
    ]


def _place_group_legend(ax: plt.Axes, groups: List[str], palette: Dict[str, str], style: dict) -> None:
    handles = _legend_handles_from_groups(groups, palette)
    ax.legend(
        handles=handles,
        title="Group",
        loc="upper left",
        bbox_to_anchor=(1.01, 1.0),
        borderaxespad=0.0,
        frameon=False,
        fontsize=style["legend_size"],
        title_fontsize=style["legend_size"],
    )


def _group_summary(per_sample: pd.DataFrame, ordered_groups: List[str]) -> pd.DataFrame:
    rows = []
    for group in ordered_groups:
        vals = pd.to_numeric(
            per_sample.loc[per_sample["Group"].astype(str) == str(group), "TotalIntensity"],
            errors="coerce",
        ).dropna()
        n = int(len(vals))
        mean_val = float(vals.mean()) if n else np.nan
        std_val = float(vals.std(ddof=1)) if n >= 2 else np.nan
        stderr_val = float(std_val / np.sqrt(n)) if n >= 2 and pd.notna(std_val) else np.nan
        rsd_val = float((std_val / mean_val) * 100.0) if n >= 2 and pd.notna(std_val) and pd.notna(mean_val) and abs(mean_val) > 1e-12 else np.nan
        rows.append({
            "Group": str(group),
            "n": n,
            "average": mean_val,
            "stddev": std_val,
            "stderr": stderr_val,
            "RSD": rsd_val,
        })
    return pd.DataFrame(rows)


def _pairwise_summary(per_sample: pd.DataFrame, ordered_groups: List[str]) -> pd.DataFrame:
    rows = []
    for g1, g2 in combinations(ordered_groups, 2):
        vals1 = pd.to_numeric(
            per_sample.loc[per_sample["Group"].astype(str) == str(g1), "TotalIntensity"],
            errors="coerce",
        ).dropna()
        vals2 = pd.to_numeric(
            per_sample.loc[per_sample["Group"].astype(str) == str(g2), "TotalIntensity"],
            errors="coerce",
        ).dropna()
        mean1 = float(vals1.mean()) if len(vals1) else np.nan
        mean2 = float(vals2.mean()) if len(vals2) else np.nan
        pval = _welch_pvalue(vals1, vals2)

        fold_change_12 = float(mean1 / mean2) if pd.notna(mean1) and pd.notna(mean2) and abs(mean2) > 1e-12 else np.nan
        log2_fc_12 = float(np.log2(fold_change_12)) if pd.notna(fold_change_12) and fold_change_12 > 0 else np.nan
        rows.append({
            "Group1": str(g1),
            "Group2": str(g2),
            "n_Group1": int(len(vals1)),
            "n_Group2": int(len(vals2)),
            "average_Group1": mean1,
            "average_Group2": mean2,
            "Fold Change (Group1/Group2)": fold_change_12,
            "log2(Fold Change)": log2_fc_12,
            "Welch p-value": pval,
        })

        fold_change_21 = float(mean2 / mean1) if pd.notna(mean1) and pd.notna(mean2) and abs(mean1) > 1e-12 else np.nan
        log2_fc_21 = float(np.log2(fold_change_21)) if pd.notna(fold_change_21) and fold_change_21 > 0 else np.nan
        rows.append({
            "Group1": str(g2),
            "Group2": str(g1),
            "n_Group1": int(len(vals2)),
            "n_Group2": int(len(vals1)),
            "average_Group1": mean2,
            "average_Group2": mean1,
            "Fold Change (Group1/Group2)": fold_change_21,
            "log2(Fold Change)": log2_fc_21,
            "Welch p-value": pval,
        })
    return pd.DataFrame(rows)


def _write_workbook(
    out_path: Path,
    group_summary: pd.DataFrame,
    pairwise_summary: pd.DataFrame,
    per_sample: pd.DataFrame,
) -> None:
    wb = Workbook()

    ws = wb.active
    ws.title = "Group summary"
    ws.append(group_summary.columns.tolist())
    for row in group_summary.itertuples(index=False, name=None):
        ws.append(list(row))

    ws2 = wb.create_sheet("Pairwise comparisons")
    ws2.append(pairwise_summary.columns.tolist())
    for row in pairwise_summary.itertuples(index=False, name=None):
        ws2.append(list(row))

    ws3 = wb.create_sheet("Per sample totals")
    ws3.append(per_sample.columns.tolist())
    for row in per_sample.itertuples(index=False, name=None):
        ws3.append(list(row))

    for sheet in wb.worksheets:
        for cell in sheet[1]:
            cell.font = Font(bold=True)
        for col in sheet.columns:
            max_len = max(len(str(cell.value)) if cell.value is not None else 0 for cell in col)
            sheet.column_dimensions[col[0].column_letter].width = min(max(max_len + 2, 12), 28)

    wb.save(out_path)


def _plot_group_totals(
    group_summary: pd.DataFrame,
    palette: Dict[str, str],
    out_png: Path,
    out_svg: Path,
    y_label: str,
    style: dict,
) -> None:
    style = _presentation_style(style)
    plot_df = group_summary.copy()
    x = np.arange(len(plot_df))
    groups = plot_df["Group"].astype(str).tolist()
    colors = [palette.get(str(g), "#777777") for g in plot_df["Group"].astype(str)]
    vals = pd.to_numeric(plot_df["average"], errors="coerce").to_numpy(dtype=float)
    errs = pd.to_numeric(plot_df["stddev"], errors="coerce").to_numpy(dtype=float)
    errs = np.where(np.isfinite(errs), errs, 0.0)

    fig, ax = plt.subplots(figsize=(16, 9), facecolor="white")
    ax.set_facecolor("white")
    ymax = float(np.nanmax(vals + errs)) if len(vals) else np.nan
    ax.set_xlim(-0.55, len(plot_df) - 0.45)
    ax.bar(
        x,
        vals,
        yerr=errs,
        color=colors,
        width=0.34,
        edgecolor="none",
        zorder=2,
        error_kw={"elinewidth": 1.2, "ecolor": "#444444", "capsize": 3, "capthick": 1.2, "zorder": 3},
    )
    ax.set_xticks(x)
    ax.set_xticklabels(_wrap_group_labels(groups), rotation=0, ha="center", fontsize=style["tick_size"])
    ax.set_ylabel(y_label, fontsize=style["label_size"], labelpad=12)
    ax.set_title("Total intensities by group (mean +/- SD)", fontsize=style["title_size"], pad=12)
    ax.tick_params(axis="y", labelsize=style["tick_size"])
    ax.set_ylim(0, ymax * 1.14 if np.isfinite(ymax) and ymax > 0 else 1.0)
    _style_axes(ax, style)
    _add_y_axis_fade(ax, ax.get_ylim()[1], ax.get_xlim()[0])
    _place_group_legend(ax, groups, palette, style)
    fig.subplots_adjust(left=0.12, right=0.78, bottom=0.24, top=0.88)
    fig.savefig(out_png, dpi=style["dpi"], facecolor="white")
    fig.savefig(out_svg, facecolor="white")
    plt.close(fig)


def _plot_group_rsd(
    group_summary: pd.DataFrame,
    palette: Dict[str, str],
    out_png: Path,
    out_svg: Path,
    y_label: str,
    style: dict,
) -> None:
    style = _presentation_style(style)
    plot_df = group_summary.copy()
    x = np.arange(len(plot_df))
    groups = plot_df["Group"].astype(str).tolist()
    colors = [palette.get(str(g), "#777777") for g in plot_df["Group"].astype(str)]
    vals = pd.to_numeric(plot_df["RSD"], errors="coerce").to_numpy(dtype=float)

    fig, ax = plt.subplots(figsize=(16, 9), facecolor="white")
    ax.set_facecolor("white")
    ymax = float(np.nanmax(vals)) if len(vals) else np.nan
    ax.set_xlim(-0.55, len(plot_df) - 0.45)
    ax.bar(x, vals, color=colors, width=0.34, edgecolor="none", zorder=2)
    ax.set_xticks(x)
    ax.set_xticklabels(_wrap_group_labels(groups), rotation=0, ha="center", fontsize=style["tick_size"])
    ax.set_ylabel(f"RSD (%) for {y_label.lower()}", fontsize=style["label_size"], labelpad=12)
    ax.set_title("RSD of total intensities by group", fontsize=style["title_size"], pad=12)
    ax.tick_params(axis="y", labelsize=style["tick_size"])
    ax.set_ylim(0, ymax * 1.14 if np.isfinite(ymax) and ymax > 0 else 1.0)
    _style_axes(ax, style)
    _add_y_axis_fade(ax, ax.get_ylim()[1], ax.get_xlim()[0])
    _place_group_legend(ax, groups, palette, style)
    fig.subplots_adjust(left=0.12, right=0.78, bottom=0.24, top=0.88)
    fig.savefig(out_png, dpi=style["dpi"], facecolor="white")
    fig.savefig(out_svg, facecolor="white")
    plt.close(fig)


def _plot_sample_totals(
    per_sample: pd.DataFrame,
    palette: Dict[str, str],
    out_png: Path,
    out_svg: Path,
    y_label: str,
    style: dict,
) -> None:
    style = _presentation_style(style)
    plot_df = per_sample.copy().reset_index(drop=True)
    x = np.arange(len(plot_df))
    colors = [palette.get(str(g), "#777777") for g in plot_df["Group"].astype(str)]
    vals = pd.to_numeric(plot_df["TotalIntensity"], errors="coerce").to_numpy(dtype=float)
    groups = list(dict.fromkeys(plot_df["Group"].astype(str).tolist()))

    fig, ax = plt.subplots(figsize=(16, 9), facecolor="white")
    ax.set_facecolor("white")
    ymax = float(np.nanmax(vals)) if len(vals) else np.nan
    ax.set_xlim(-0.55, len(plot_df) - 0.45)
    ax.bar(x, vals, color=colors, width=0.62, edgecolor="none", zorder=2)
    ax.set_xticks(x)
    ax.set_xticklabels(plot_df["Sample"].astype(str), rotation=90, ha="center", fontsize=max(style["tick_size"] - 2, 12))
    ax.set_ylabel(y_label, fontsize=style["label_size"], labelpad=12)
    ax.set_title("Total intensities per sample", fontsize=style["title_size"], pad=12)
    ax.tick_params(axis="y", labelsize=style["tick_size"])
    ax.set_ylim(0, ymax * 1.14 if np.isfinite(ymax) and ymax > 0 else 1.0)
    _style_axes(ax, style)
    _add_y_axis_fade(ax, ax.get_ylim()[1], ax.get_xlim()[0])
    _place_group_legend(ax, groups, palette, style)
    fig.subplots_adjust(left=0.12, right=0.78, bottom=0.32, top=0.88)
    fig.savefig(out_png, dpi=style["dpi"], facecolor="white")
    fig.savefig(out_svg, facecolor="white")
    plt.close(fig)


def run_from_stats(
    file_path: str,
    group_file: Optional[str],
    save_dir: str,
    group_order: Optional[List[str]] = None,
    group_colors: Optional[dict] = None,
    dataset_label: Optional[str] = None,
    dpi: int = 100,
    publication_theme: bool = False,
) -> Dict[str, str]:
    out_dir = Path(prepare_output_dir(save_dir))
    style = _presentation_style(get_figure_style(publication_theme=publication_theme, dpi=dpi))

    print("[TotalIntensity] Running total intensity analysis...", flush=True)
    X, y, _ = load_dataset(file_path, group_file)
    if X.empty or y.empty:
        raise ValueError("Dataset appears empty or malformed.")

    per_sample = pd.DataFrame({
        "Sample": X.index.astype(str),
        "Group": y.astype(str).reindex(X.index).values,
        "TotalIntensity": pd.to_numeric(X.sum(axis=1), errors="coerce").to_numpy(dtype=float),
    })
    ordered_groups, palette = _build_palette(
        per_sample["Group"].astype(str),
        group_colors=group_colors,
        group_order=_order_groups(pd.unique(per_sample["Group"].astype(str)).tolist(), group_order),
    )
    per_sample["Group"] = pd.Categorical(per_sample["Group"].astype(str), categories=ordered_groups, ordered=True)
    per_sample = per_sample.sort_values(["Group", "Sample"], kind="stable").reset_index(drop=True)

    group_summary = _group_summary(per_sample, ordered_groups)
    pairwise_summary = _pairwise_summary(per_sample, ordered_groups)

    per_sample_csv = out_dir / "total_intensity_per_sample.csv"
    group_summary_csv = out_dir / "total_intensity_group_summary.csv"
    pairwise_csv = out_dir / "total_intensity_pairwise_comparisons.csv"
    workbook_path = out_dir / "total_intensity_summary.xlsx"

    per_sample.to_csv(per_sample_csv, index=False)
    group_summary.to_csv(group_summary_csv, index=False)
    pairwise_summary.to_csv(pairwise_csv, index=False)
    _write_workbook(workbook_path, group_summary, pairwise_summary, per_sample)

    y_label = _total_intensity_label(dataset_label, file_path)
    totals_png = out_dir / "total_intensity_by_group.png"
    totals_svg = out_dir / "total_intensity_by_group.svg"
    rsd_png = out_dir / "total_intensity_rsd_by_group.png"
    rsd_svg = out_dir / "total_intensity_rsd_by_group.svg"
    sample_png = out_dir / "total_intensity_per_sample.png"
    sample_svg = out_dir / "total_intensity_per_sample.svg"

    _plot_group_totals(group_summary, palette, totals_png, totals_svg, y_label, style)
    _plot_group_rsd(group_summary, palette, rsd_png, rsd_svg, y_label, style)
    _plot_sample_totals(per_sample, palette, sample_png, sample_svg, y_label, style)

    return {
        "out_dir": str(out_dir),
        "per_sample_csv": str(per_sample_csv),
        "group_summary_csv": str(group_summary_csv),
        "pairwise_csv": str(pairwise_csv),
        "workbook": str(workbook_path),
        "group_totals_png": str(totals_png),
        "group_totals_svg": str(totals_svg),
        "group_rsd_png": str(rsd_png),
        "group_rsd_svg": str(rsd_svg),
        "sample_totals_png": str(sample_png),
        "sample_totals_svg": str(sample_svg),
    }
