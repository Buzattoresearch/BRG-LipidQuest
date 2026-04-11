# TODO: check if sample type is being passed from the GUI
# TODO: Annotation type and headgroup are not showing in the Excel and CSV files
# TODO: confirm that the Summary_Volcano file is actually correct

# ------------------------------------------------------------
# Volcano analysis for standardized Lipid workflow (LipidQuest)
# - Loads via Stats.utils.load_dataset(file_path, group_file)
# - Pairwise group comparisons (t-test or Mann–Whitney)
# - FDR correction (Benjamini–Hochberg)
# - Volcano plots (PNG+SVG), per-comparison CSVs
# - Summary CSV + structured Excel
# - Class bar plots + two bubble-plot designs
# - Saves all results under save_dir/VolcanoFDR/
# ------------------------------------------------------------

'''
STATISTICAL LOGIC USED IN THIS VOLCANO ROUTINE

This module performs pairwise lipid comparisons between two biological groups and reports
fold change, raw p-value, FDR-adjusted p-value, assumption checks, test used, and effect size.

Why this is needed:
Lipidomics datasets often have small sample sizes, missing values, unequal variances, and
strong correlation among lipids within the same class. Under these conditions, raw p-values
alone are unstable and may fail to capture biologically meaningful changes. This routine
therefore reports significance metrics and effect-size metrics, and can switch between
parametric and nonparametric tests on a per-feature basis.

Main statistics used:

1. Fold Change
   Fold change is the ratio of the central tendency in group 1 relative to group 2.
   Here, fold change is calculated from group medians, not means, to reduce sensitivity
   to outliers. A small per-feature pseudocount is added to avoid division by zero.
   log2(Fold Change) is used for volcano plotting because it is symmetric around zero:
   positive values indicate higher abundance in group 1, and negative values indicate
   lower abundance in group 1.

2. Shapiro-Wilk normality test
   Shapiro-Wilk tests whether the values within one group are consistent with a normal
   distribution. It is run separately for each group.
   - Null hypothesis: the data are normally distributed.
   - A large p-value suggests no evidence against normality.
   - A small p-value suggests deviation from normality.
   Important limitation: with very small n, Shapiro-Wilk has low power and may fail to
   detect non-normality. Therefore, these results are reported as assumption diagnostics,
   not as a definitive statement about the data distribution.

3. Levene test for equality of variance
   Levene tests whether the two groups have similar variance.
   - Null hypothesis: the group variances are equal.
   - A large p-value suggests no evidence against equal variance.
   - A small p-value suggests unequal variance.
   The median-centered version is used because it is more robust than the classical
   variance tests when the data are not perfectly normal.

4. Student t-test
   Student's t-test compares group means under the assumptions that:
   - the data are approximately normal
   - the group variances are approximately equal
   This test is available for manual use, but is not the default in this routine.

5. Welch t-test
   Welch's t-test also compares group means, but does not assume equal variance between
   groups. It is generally safer than Student's t-test for omics data and is used here
   as the default test.
   In this routine:
   - test_type = "welch" or "parametric" uses Welch's t-test
   - test_type = "auto" also defaults to Welch's t-test for all features

6. Mann-Whitney U test
   Mann-Whitney is a nonparametric test that compares the rank distributions of the two groups.
   It does not require normally distributed data and can be used as a robustness check
   when distributions look strongly non-normal, zero-inflated, or outlier-driven.
   Important limitation: Mann-Whitney is often described as a test of medians, but that is
   only strictly true under specific shape assumptions. More generally, it tests whether one
   group tends to have larger values than the other.

7. Automatic test selection
   When test_type = "auto", the routine currently defaults to Welch's t-test for every lipid.
   Shapiro-Wilk and Levene are still reported in the output tables as assumption diagnostics,
   but they do not change the primary test selection.
   Mann-Whitney results are also reported as a robustness check, including the
   Mann-Whitney p-value, rank-biserial direction, and whether that direction agrees
   with the fold-change direction. Nonparametric support is only evaluated for lipids
   that pass the primary Welch raw p-value or FDR threshold; otherwise it is marked
   as not evaluated.

8. Raw p-value
   The raw p-value is the probability, under the null hypothesis of no group difference,
   of observing data at least as extreme as those obtained. Raw p-values are reported for
   each lipid, but should not be interpreted alone when many lipids are tested at once.

9. False Discovery Rate correction
   Raw p-values are adjusted using the Benjamini-Hochberg procedure to control the false
   discovery rate (FDR) across all tested lipids in a comparison.
   - FDR-adjusted p-values reduce the number of false positives expected from multiple testing.
   - This procedure assumes independence or weak positive dependence among tests.
   Lipidomics data often violate this assumption because lipids within the same class are
   biologically and analytically correlated. Therefore, FDR can be conservative and may
   remove biologically meaningful signals, especially when n is small.

10. Effect size
   Effect size quantifies the magnitude of the difference between groups, independent of
   sample size. This is important because a biologically strong shift may fail to reach
   statistical significance in small studies.

   The main effect-size column uses Hedges' g for all lipids. In addition, a
   Mann-Whitney rank-biserial correlation is reported separately as part of the
   nonparametric robustness output.

   a. Hedges' g
      Hedges' g is a small-sample corrected standardized mean difference.
      It is used here as the primary effect size for all lipids.
      - Positive values indicate that group 1 is higher than group 2.
      - Negative values indicate that group 1 is lower than group 2.
      Approximate interpretation:
      |g| ~ 0.2 small
      |g| ~ 0.5 moderate
      |g| ~ 0.8 large

   b. Rank-biserial correlation
      Rank-biserial correlation is derived from the Mann-Whitney U statistic and is
      reported as a separate nonparametric robustness metric.
      - Positive values indicate that group 1 tends to have larger values than group 2.
      - Negative values indicate that group 1 tends to have smaller values than group 2.
      Approximate interpretation:
      |r| ~ 0.1 small
      |r| ~ 0.3 moderate
      |r| ~ 0.5 large

11. Large-effect flag
   A lipid is flagged as "Large Effect" even if it does not pass FDR when:
   - |Hedges' g| >= 0.8
   This flag is intended to highlight potentially meaningful biological shifts that may
   be missed by strict multiple-testing thresholds in small datasets.

12. Volcano significance
   In the volcano output, a feature is labeled as significantly increased or decreased only if:
   - the fold-change threshold is passed
   - the raw p-value threshold is passed
   - the FDR threshold is passed
   Features that fail any of these criteria are labeled as not significant, even if they
   show a moderate or large effect size.

Important interpretation notes:
- Small n reduces the reliability of normality tests and variance estimates.
- FDR correction can be overly conservative in lipidomics because lipids are highly correlated.
- A non-significant volcano plot does not imply absence of biology. The data may still contain
  consistent class-level remodeling or moderate effect sizes that do not survive multiple testing.
- Effect size should always be interpreted together with fold change, missingness, reproducibility,
  and biological plausibility.
'''

import os
import re
import warnings
from pathlib import Path
from typing import Optional, Union

import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
import seaborn as sns

from scipy.stats import ttest_ind, mannwhitneyu, shapiro, levene
from pandas import IndexSlice
from matplotlib.colors import LinearSegmentedColormap, Normalize, TwoSlopeNorm
from matplotlib.lines import Line2D
from matplotlib.patches import Patch
from statsmodels.stats.multitest import multipletests
from Stats.figure_style import VOLCANO_DOWN, VOLCANO_NS, VOLCANO_UP, get_figure_style
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from Stats.utils import load_dataset, prepare_output_dir
from Stats.utils import _CLASS_ORDER, _CLASS_ORDER_BACTERIA, _CLASS_ORDER_MAMMALIAN, _CLASS_GROUP_MAP

import warnings
warnings.simplefilter("ignore", pd.errors.PerformanceWarning)

import matplotlib as mpl
mpl.rcParams["font.family"] = "Arial"
mpl.rcParams["font.sans-serif"] = ["Arial", "Liberation Sans", "DejaVu Sans"]
mpl.rcParams["mathtext.default"] = "regular" 

plt.rcParams["font.size"] = 14
warnings.filterwarnings("ignore", category=FutureWarning)
warnings.filterwarnings("ignore", message="Glyph .* missing from font.*")

plt.ioff()

# ==========================================================
# Utilities (general)
# ==========================================================
def _sanitize_filename(s: str) -> str:
    return re.sub(r'[<>:."/\\|?*]', "_", str(s))

def _normalize_windows_save_path(path: Path) -> str:
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    path_str = str(path.resolve())
    if os.name == "nt" and not path_str.startswith("\\\\?\\"):
        path_str = f"\\\\?\\{path_str}"
    return path_str

def _save_figure(fig, path: Path, **kwargs):
    fig.savefig(_normalize_windows_save_path(path), **kwargs)

def _save_current_figure(path: Path, **kwargs):
    plt.savefig(_normalize_windows_save_path(path), **kwargs)

def _add_jitter(values, jitter_strength=0.025):
    j = values + np.random.uniform(-jitter_strength, jitter_strength, size=len(values))
    return np.maximum(0.0, j)   # clip at 0

def _ensure_dir(p: Path) -> Path:
    p.mkdir(parents=True, exist_ok=True)
    return p

def _safe_shapiro(x: Union[pd.Series, np.ndarray]) -> float:
    """
    Return Shapiro-Wilk p-value.
    Returns NaN when the test is not valid or not informative.
    """
    x = pd.Series(x).replace([np.inf, -np.inf], np.nan).dropna().astype(float)
    if len(x) < 3:
        return np.nan
    if x.nunique(dropna=True) < 3:
        return np.nan
    try:
        return float(shapiro(x)[1])
    except Exception:
        return np.nan

def _safe_levene(x1: Union[pd.Series, np.ndarray], x2: Union[pd.Series, np.ndarray]) -> float:
    """
    Return Levene p-value using median-centered version.
    Returns NaN when the test is not valid.
    """
    x1 = pd.Series(x1).replace([np.inf, -np.inf], np.nan).dropna().astype(float)
    x2 = pd.Series(x2).replace([np.inf, -np.inf], np.nan).dropna().astype(float)
    if len(x1) < 2 or len(x2) < 2:
        return np.nan
    try:
        return float(levene(x1, x2, center="median")[1])
    except Exception:
        return np.nan

def _hedges_g(x1: Union[pd.Series, np.ndarray], x2: Union[pd.Series, np.ndarray]) -> float:
    """
    Small-sample corrected standardized mean difference.
    Positive values mean group1 > group2.
    """
    x1 = pd.Series(x1).replace([np.inf, -np.inf], np.nan).dropna().astype(float)
    x2 = pd.Series(x2).replace([np.inf, -np.inf], np.nan).dropna().astype(float)

    n1, n2 = len(x1), len(x2)
    if n1 < 2 or n2 < 2:
        return np.nan

    s1 = float(np.std(x1, ddof=1))
    s2 = float(np.std(x2, ddof=1))
    if not np.isfinite(s1) or not np.isfinite(s2):
        return np.nan

    pooled_num = ((n1 - 1) * (s1 ** 2)) + ((n2 - 1) * (s2 ** 2))
    pooled_den = n1 + n2 - 2
    if pooled_den <= 0:
        return np.nan

    s_pooled = np.sqrt(pooled_num / pooled_den)
    if s_pooled < 1e-12:
        return 0.0

    d = (float(np.mean(x1)) - float(np.mean(x2))) / s_pooled

    # Hedges correction
    N = n1 + n2
    if N <= 3:
        return np.nan
    J = 1.0 - (3.0 / (4.0 * N - 9.0))
    return float(d * J)

def _rank_biserial_from_mwu(x1: Union[pd.Series, np.ndarray], x2: Union[pd.Series, np.ndarray]) -> float:
    """
    Rank-biserial correlation derived from Mann-Whitney U.
    Positive values mean group1 tends to be larger than group2.
    """
    x1 = pd.Series(x1).replace([np.inf, -np.inf], np.nan).dropna().astype(float)
    x2 = pd.Series(x2).replace([np.inf, -np.inf], np.nan).dropna().astype(float)

    n1, n2 = len(x1), len(x2)
    if n1 < 1 or n2 < 1:
        return np.nan

    try:
        u_stat, _ = mannwhitneyu(
            x1,
            x2,
            alternative="two-sided",
            method="asymptotic",
            use_continuity=False,
        )
        rbc = (2.0 * u_stat / (n1 * n2)) - 1.0
        return float(rbc)
    except Exception:
        return np.nan

def _choose_test(x1, x2, test_type: str, alpha_assumption: float = 0.05):
    """
    Returns:
        test_used, p_value, normality_p_g1, normality_p_g2, variance_p
    """
    x1 = pd.Series(x1).replace([np.inf, -np.inf], np.nan).dropna().astype(float)
    x2 = pd.Series(x2).replace([np.inf, -np.inf], np.nan).dropna().astype(float)

    n1, n2 = len(x1), len(x2)

    shapiro_p1 = _safe_shapiro(x1)
    shapiro_p2 = _safe_shapiro(x2)
    levene_p = _safe_levene(x1, x2)

    if n1 < 2 or n2 < 2:
        return "insufficient_n", 1.0, shapiro_p1, shapiro_p2, levene_p

    if (np.std(x1, ddof=1) < 1e-12) and (np.std(x2, ddof=1) < 1e-12):
        return "constant", 1.0, shapiro_p1, shapiro_p2, levene_p

    normalized_test_type = str(test_type or "auto").strip().lower()

    try:
        if normalized_test_type in {"non-parametric", "nonparametric", "mann-whitney", "mannwhitney"}:
            _, p = mannwhitneyu(
                x1,
                x2,
                alternative="two-sided",
                method="asymptotic",
                use_continuity=False,
            )
            return "mannwhitney", float(p), shapiro_p1, shapiro_p2, levene_p

        if normalized_test_type in {"parametric", "welch", "welch_t", "welch's", "welchs"}:
            _, p = ttest_ind(x1, x2, equal_var=False, nan_policy="omit")
            return "welch_t", float(p), shapiro_p1, shapiro_p2, levene_p

        if normalized_test_type in {"student", "student_t", "student's", "students"}:
            _, p = ttest_ind(x1, x2, equal_var=True, nan_policy="omit")
            return "student_t", float(p), shapiro_p1, shapiro_p2, levene_p

        if normalized_test_type == "auto":
            _, p = ttest_ind(x1, x2, equal_var=False, nan_policy="omit")
            return "welch_t", float(p), shapiro_p1, shapiro_p2, levene_p

    except Exception:
        pass

    return "failed", 1.0, shapiro_p1, shapiro_p2, levene_p


def _compute_all_test_pvalues(x1, x2, alpha_assumption: float = 0.05):
    auto_test_used, auto_p, shapiro_p1, shapiro_p2, levene_p = _choose_test(
        x1, x2,
        test_type="auto",
        alpha_assumption=alpha_assumption,
    )
    _, student_p, _, _, _ = _choose_test(
        x1, x2,
        test_type="student",
        alpha_assumption=alpha_assumption,
    )
    _, welch_p, _, _, _ = _choose_test(
        x1, x2,
        test_type="welch",
        alpha_assumption=alpha_assumption,
    )
    _, mannwhitney_p, _, _, _ = _choose_test(
        x1, x2,
        test_type="mann-whitney",
        alpha_assumption=alpha_assumption,
    )
    return {
        "auto_test_used": auto_test_used,
        "auto_p": float(auto_p),
        "student_p": float(student_p),
        "welch_p": float(welch_p),
        "mannwhitney_p": float(mannwhitney_p),
        "shapiro_p1": shapiro_p1,
        "shapiro_p2": shapiro_p2,
        "levene_p": levene_p,
    }


def _pick_column_ci(df: pd.DataFrame, candidates: list[str]) -> Optional[str]:
    if df is None or df.empty:
        return None
    norm = {str(c).strip().lower(): str(c) for c in df.columns}
    for cand in candidates:
        key = str(cand).strip().lower()
        if key in norm:
            return norm[key]
    return None


def _parse_total_carbons_from_text(s: str) -> float:
    if s is None:
        return np.nan
    m = re.search(r"(\d+)\s*:\s*(\d+)", str(s))
    return float(int(m.group(1))) if m else np.nan


def _parse_total_double_bonds_from_text(s: str) -> float:
    if s is None:
        return np.nan
    m = re.search(r"(\d+)\s*:\s*(\d+)", str(s))
    return float(int(m.group(2))) if m else np.nan


def _append_species_coherence_metrics(df: pd.DataFrame) -> pd.DataFrame:
    if df is None or df.empty:
        return df

    out = df.copy()
    out["Lipid Class"] = (
        out.get("Lipid Class", pd.Series(index=out.index, dtype=object))
        .astype(str)
        .str.strip()
        .replace({"": "Unknown", "nan": "Unknown", "None": "Unknown"})
    )
    out.loc[out["Lipid Class"].eq("Unknown"), "Lipid Class"] = out.loc[
        out["Lipid Class"].eq("Unknown"), "UniqueID"
    ].apply(_extract_Class)

    carb = pd.to_numeric(out.get("Number of carbons in fatty acyls"), errors="coerce")
    dbe = pd.to_numeric(out.get("Double bond equivalents"), errors="coerce")

    if "Annotation" in out.columns:
        carb = carb.where(carb.notna(), out["Annotation"].apply(_parse_total_carbons_from_text))
        dbe = dbe.where(dbe.notna(), out["Annotation"].apply(_parse_total_double_bonds_from_text))
    carb = carb.where(carb.notna(), out["UniqueID"].apply(_parse_total_carbons_from_text))
    dbe = dbe.where(dbe.notna(), out["UniqueID"].apply(_parse_total_double_bonds_from_text))

    out["Number of carbons in fatty acyls"] = carb
    out["Double bond equivalents"] = dbe
    out["_direction_sign"] = np.sign(pd.to_numeric(out["log2(Fold Change)"], errors="coerce")).fillna(0.0).astype(int)

    class_n = []
    class_same_n = []
    class_same_frac = []
    class_median_fc = []
    neighbor_n = []
    neighbor_same_n = []
    neighbor_same_frac = []
    neighbor_median_fc = []

    for idx in out.index:
        lipid_class = out.at[idx, "Lipid Class"]
        lipid_sign = out.at[idx, "_direction_sign"]
        lipid_carb = out.at[idx, "Number of carbons in fatty acyls"]
        lipid_dbe = out.at[idx, "Double bond equivalents"]

        class_mask = out["Lipid Class"].eq(lipid_class)
        class_df = out.loc[class_mask]
        class_count = int(len(class_df))
        class_same = int((class_df["_direction_sign"] == lipid_sign).sum()) if class_count else 0

        class_n.append(class_count)
        class_same_n.append(class_same)
        class_same_frac.append(float(class_same / class_count) if class_count > 0 else np.nan)
        class_median_fc.append(float(pd.to_numeric(class_df["log2(Fold Change)"], errors="coerce").median()) if class_count > 0 else np.nan)

        if pd.notna(lipid_carb) and pd.notna(lipid_dbe):
            nbr_mask = (
                out["Lipid Class"].eq(lipid_class)
                & pd.to_numeric(out["Number of carbons in fatty acyls"], errors="coerce").sub(float(lipid_carb)).abs().le(2)
                & pd.to_numeric(out["Double bond equivalents"], errors="coerce").sub(float(lipid_dbe)).abs().le(1)
            )
            nbr_df = out.loc[nbr_mask]
            nbr_count = int(len(nbr_df))
            nbr_same = int((nbr_df["_direction_sign"] == lipid_sign).sum()) if nbr_count else 0
            neighbor_n.append(nbr_count)
            neighbor_same_n.append(nbr_same)
            neighbor_same_frac.append(float(nbr_same / nbr_count) if nbr_count > 0 else np.nan)
            neighbor_median_fc.append(float(pd.to_numeric(nbr_df["log2(Fold Change)"], errors="coerce").median()) if nbr_count > 0 else np.nan)
        else:
            neighbor_n.append(0)
            neighbor_same_n.append(0)
            neighbor_same_frac.append(np.nan)
            neighbor_median_fc.append(np.nan)

    out["Class_n"] = class_n
    out["Class_same_direction_n"] = class_same_n
    out["Class_same_direction_frac"] = class_same_frac
    out["Class_median_log2FC"] = class_median_fc
    out["Neighbor_n"] = neighbor_n
    out["Neighbor_same_direction_n"] = neighbor_same_n
    out["Neighbor_same_direction_frac"] = neighbor_same_frac
    out["Neighbor_median_log2FC"] = neighbor_median_fc
    out["Class_consistency_flag"] = np.where(
        (pd.to_numeric(out["Class_n"], errors="coerce") >= 3)
        & (pd.to_numeric(out["Class_same_direction_frac"], errors="coerce") >= 0.70),
        "Yes",
        "No",
    )
    out["Neighbor_consistency_flag"] = np.where(
        (pd.to_numeric(out["Neighbor_n"], errors="coerce") >= 2)
        & (pd.to_numeric(out["Neighbor_same_direction_frac"], errors="coerce") >= 0.70),
        "Yes",
        "No",
    )
    out = out.drop(columns=["_direction_sign"], errors="ignore")
    return out


def _save_class_enrichment_summary(csv_dir: Path, output_csv: Path) -> None:
    files = sorted([f for f in os.listdir(csv_dir) if f.endswith("_FDR.csv")])
    if not files:
        return

    rows = []
    for fname in files:
        comparison = fname.replace("_FDR.csv", "")
        df = pd.read_csv(csv_dir / fname)
        if df.empty or "UniqueID" not in df.columns:
            continue

        if "Lipid Class" not in df.columns:
            df["Lipid Class"] = df["UniqueID"].apply(_extract_Class)
        df["Lipid Class"] = (
            df["Lipid Class"].astype(str).str.strip().replace({"": "Unknown", "nan": "Unknown"})
        )

        abs_effect = pd.to_numeric(
            df.get("Effect Size (Hedge's g)", pd.Series(np.nan, index=df.index)),
            errors="coerce",
        ).abs()
        effect_thresh = abs_effect.quantile(0.90) if abs_effect.notna().any() else np.nan
        raw_p = pd.to_numeric(df.get("p-value"), errors="coerce")
        abs_log2fc = pd.to_numeric(df.get("log2(Fold Change)"), errors="coerce").abs()

        class_groups = df.groupby("Lipid Class", dropna=False)
        for lipid_class, class_df in class_groups:
            class_abs_effect = pd.to_numeric(
                class_df.get("Effect Size (Hedge's g)", pd.Series(np.nan, index=class_df.index)),
                errors="coerce",
            ).abs()
            class_log2fc = pd.to_numeric(class_df.get("log2(Fold Change)"), errors="coerce")
            class_raw_p = pd.to_numeric(class_df.get("p-value"), errors="coerce")
            up_n = int(class_log2fc.gt(0).sum())
            down_n = int(class_log2fc.lt(0).sum())
            total_n = int(len(class_df))
            rows.append({
                "Comparison": comparison,
                "Lipid Class": lipid_class,
                "Total detected lipids": total_n,
                "Lipids with raw p < 0.05": int(class_raw_p.lt(0.05).sum()),
                "Lipids in top 10% by absolute effect size": int(
                    class_abs_effect.ge(effect_thresh).sum()
                ) if np.isfinite(effect_thresh) else 0,
                "Median log2FC": float(class_log2fc.median()),
                "Mean log2FC": float(class_log2fc.mean()),
                "Increased_n": up_n,
                "Decreased_n": down_n,
                "Increased_frac": float(up_n / total_n) if total_n > 0 else np.nan,
                "Decreased_frac": float(down_n / total_n) if total_n > 0 else np.nan,
                "Same_direction_frac": float(max(up_n, down_n) / total_n) if total_n > 0 else np.nan,
                "Significant_up_n": int((class_raw_p.lt(0.05) & class_log2fc.gt(0)).sum()),
                "Significant_down_n": int((class_raw_p.lt(0.05) & class_log2fc.lt(0)).sum()),
                "Median absolute log2FC": float(pd.to_numeric(class_df.get("log2(Fold Change)"), errors="coerce").abs().median()),
                "Median absolute effect size": float(class_abs_effect.median()),
            })

    if rows:
        pd.DataFrame(rows).to_csv(output_csv, index=False, encoding="utf-8-sig")

# ==========================================================
# Volcano core computation + plotting
# ==========================================================
def _compute_volcano(g1_name, g2_name, X, y, meta_lookup,
                     method, test_type, p_thresh, fdr_thresh, fc_thresh,
                     assumption_alpha=0.05):
    """Perform one pairwise comparison and return volcano DataFrame."""
    # --- Select groups and common features ---
    y_str = y.astype(str)
    group1 = X.loc[y_str == str(g1_name)]
    group2 = X.loc[y_str == str(g2_name)]

    common = group1.columns.intersection(group2.columns)
    group1 = group1[common]
    group2 = group2[common]

    # --- Remove features that are entirely empty in BOTH groups ---
    nonzero_any = (group1.notna().any(axis=0) & group2.notna().any(axis=0))
    group1 = group1.loc[:, nonzero_any]
    group2 = group2.loc[:, nonzero_any]
    valid_UniqueIDs = list(group1.columns)

    normalized_test_type = str(test_type or "auto").strip().lower()

    feature_records = []
    for UniqueID in valid_UniqueIDs:
        x1 = group1[UniqueID].replace([np.inf, -np.inf], np.nan).dropna()
        x2 = group2[UniqueID].replace([np.inf, -np.inf], np.nan).dropna()

        n1 = int(len(x1))
        n2 = int(len(x2))

        all_test_p = _compute_all_test_pvalues(x1, x2, alpha_assumption=assumption_alpha)
        shapiro_p1 = all_test_p["shapiro_p1"]
        shapiro_p2 = all_test_p["shapiro_p2"]
        levene_p = all_test_p["levene_p"]

        feature_records.append({
            "UniqueID": UniqueID,
            "x1": x1,
            "x2": x2,
            "n1": n1,
            "n2": n2,
            "shapiro_p1": shapiro_p1,
            "shapiro_p2": shapiro_p2,
            "levene_p": levene_p,
            "auto_test_used": all_test_p["auto_test_used"],
            "auto_p": float(all_test_p["auto_p"]),
            "student_p": float(all_test_p["student_p"]),
            "welch_p": float(all_test_p["welch_p"]),
            "mannwhitney_p": float(all_test_p["mannwhitney_p"]),
        })

    # --- p-values on filtered features only ---
    pvals = []
    n1_list = []
    n2_list = []
    shapiro_p1_list = []
    shapiro_p2_list = []
    levene_p_list = []
    effect_size_list = []
    auto_test_used_list = []
    auto_pvals_list = []
    student_pvals_list = []
    welch_pvals_list = []
    mannwhitney_pvals_list = []
    mannwhitney_rbc_list = []

    for rec in feature_records:
        if normalized_test_type == "auto":
            test_used = rec["auto_test_used"]
            p = rec["auto_p"]
        else:
            test_used, p, _, _, _ = _choose_test(
                rec["x1"], rec["x2"],
                test_type=test_type,
                alpha_assumption=assumption_alpha
            )

        effect_size = _hedges_g(rec["x1"], rec["x2"])
        mannwhitney_rbc = _rank_biserial_from_mwu(rec["x1"], rec["x2"])

        pvals.append(float(p))
        n1_list.append(rec["n1"])
        n2_list.append(rec["n2"])
        shapiro_p1_list.append(rec["shapiro_p1"])
        shapiro_p2_list.append(rec["shapiro_p2"])
        levene_p_list.append(rec["levene_p"])
        effect_size_list.append(effect_size)
        auto_test_used_list.append(rec["auto_test_used"])
        auto_pvals_list.append(rec["auto_p"])
        student_pvals_list.append(rec["student_p"])
        welch_pvals_list.append(rec["welch_p"])
        mannwhitney_pvals_list.append(rec["mannwhitney_p"])
        mannwhitney_rbc_list.append(mannwhitney_rbc)

    pvals = np.asarray(pvals, dtype=float)
    fdr_corrected = multipletests(pvals, alpha=0.05, method=method)[1]

    # --- Robust FC using medians + per-feature pseudocount ---
    g1_med = group1.replace(0, np.nan).median()
    g2_med = group2.replace(0, np.nan).median()

    pooled_med = pd.concat([group1, group2]).replace(0, np.nan).median()
    eps = (0.01 * pooled_med).fillna(0) + 1e-9  # 1% of pooled nonzero median, with floor

    g1m = (g1_med.fillna(0) + eps)
    g2m = (g2_med.fillna(0) + eps)

    fc = (g1m / g2m).clip(lower=1e-6, upper=1e6)
    fc_log2 = np.log2(fc)

    df = pd.DataFrame({
        "UniqueID": valid_UniqueIDs,
        "n_group1": n1_list,
        "n_group2": n2_list,
        "Fold Change": np.round(fc.values, 18),
        "log2(Fold Change)": np.round(fc_log2.values, 10),
        "p-value": np.round(pvals, 18),
        "FDR p-value": np.round(fdr_corrected, 18),
        "Auto selected test": auto_test_used_list,
        "p-value (Auto)": np.round(np.asarray(auto_pvals_list, dtype=float), 18),
        "p-value (Student)": np.round(np.asarray(student_pvals_list, dtype=float), 18),
        "p-value (Welch)": np.round(np.asarray(welch_pvals_list, dtype=float), 18),
        "p-value (Mann-Whitney)": np.round(np.asarray(mannwhitney_pvals_list, dtype=float), 18),
        "Mann-Whitney rank-biserial": mannwhitney_rbc_list,
        f"Shapiro p-value ({g1_name})": shapiro_p1_list,
        f"Shapiro p-value ({g2_name})": shapiro_p2_list,
        "Levene p-value": levene_p_list,
        "Effect Size (Hedge's g)": effect_size_list,
    })

    # Clean non-finite and compute -log10(FDR)
    df = df.replace([np.inf, -np.inf], np.nan).dropna(subset=["log2(Fold Change)", "FDR p-value"])
    df["-log10(FDR p-value)"] = -np.log10(np.clip(df["FDR p-value"], 1e-300, 1.0))
    
    df[f"Normality pass ({g1_name})"] = np.select(
        [
            df[f"Shapiro p-value ({g1_name})"].notna() & (df[f"Shapiro p-value ({g1_name})"] >= assumption_alpha),
            df[f"Shapiro p-value ({g1_name})"].notna() & (df[f"Shapiro p-value ({g1_name})"] < assumption_alpha),
        ],
        ["Yes", "No"],
        default=np.nan,
    )
    df[f"Normality pass ({g2_name})"] = np.select(
        [
            df[f"Shapiro p-value ({g2_name})"].notna() & (df[f"Shapiro p-value ({g2_name})"] >= assumption_alpha),
            df[f"Shapiro p-value ({g2_name})"].notna() & (df[f"Shapiro p-value ({g2_name})"] < assumption_alpha),
        ],
        ["Yes", "No"],
        default=np.nan,
    )
    df["Variance equal pass"] = np.where(
        df["Levene p-value"].notna(),
        df["Levene p-value"] >= assumption_alpha,
        np.nan
    )
    # Rank-biserial > 0 means group1 tends to be larger than group2, matching log2FC > 0.
    df["Mann-Whitney direction agrees"] = np.where(
        df["Mann-Whitney rank-biserial"].notna() & (np.sign(df["Mann-Whitney rank-biserial"]) == np.sign(df["log2(Fold Change)"])),
        "Yes",
        "No"
    )
    df.loc[
        df["Mann-Whitney rank-biserial"].isna() | (np.sign(df["log2(Fold Change)"]) == 0),
        "Mann-Whitney direction agrees"
    ] = "Undetermined"
    # Significance status
    log2fc_thresh = np.log2(fc_thresh)
    df["Significance"] = "Not Significant"
    df.loc[
        (df["log2(Fold Change)"] >= log2fc_thresh) &
        (df["p-value"] < p_thresh) &
        (df["FDR p-value"] < fdr_thresh),
        "Significance"
    ] = "Up"
    df.loc[
        (df["log2(Fold Change)"] <= -log2fc_thresh) &
        (df["p-value"] < p_thresh) &
        (df["FDR p-value"] < fdr_thresh),
        "Significance"
    ] = "Down"

    welch_primary_significant = (df["p-value"] < 0.05) | (df["FDR p-value"] < 0.05)
    df["Mann-Whitney robustness"] = np.select(
        [
            ~welch_primary_significant,
            welch_primary_significant & (df["p-value (Mann-Whitney)"] < p_thresh) & df["Mann-Whitney direction agrees"].eq("Yes"),
            welch_primary_significant & (df["p-value (Mann-Whitney)"] < p_thresh) & df["Mann-Whitney direction agrees"].eq("No"),
        ],
        [
            "Not evaluated (Welch p >= 0.05 and FDR >= 0.05)",
            "Supported",
            "Discordant",
        ],
        default="Not supported",
    )

    df["Large Effect"] = "No"
    df.loc[
        df["Effect Size (Hedge's g)"].abs() >= 0.8,
        "Large Effect"
    ] = "Yes"
    
    df["UniqueID"] = df["UniqueID"].astype(str).str.strip()

    # ---- Merge Annotation, Headgroup, and Class metadata ----
    if meta_lookup is not None and not meta_lookup.empty:
        df = df.merge(meta_lookup, how="left", on="UniqueID")
    else:
        df["Annotation"] = "Unknown"
        df["Annotation Type"] = "Unknown"
        df["Headgroup"] = "Unknown"
        df["Lipid Class"] = "Unknown"
        df["Number of carbons in fatty acyls"] = np.nan
        df["Double bond equivalents"] = np.nan

    df = _append_species_coherence_metrics(df)

    # ---- Reorder columns for tidy output ----
    preferred_order = [
        "UniqueID", "Annotation", "Annotation Type", "Headgroup", "Lipid Class",
        "Number of carbons in fatty acyls", "Double bond equivalents",
        "n_group1", "n_group2",
        "Fold Change", "log2(Fold Change)",
        "p-value", "FDR p-value", "-log10(FDR p-value)",
        "Auto selected test",
        "p-value (Auto)", "p-value (Student)", "p-value (Welch)", "p-value (Mann-Whitney)",
        "Mann-Whitney rank-biserial", "Mann-Whitney direction agrees", "Mann-Whitney robustness",
        f"Shapiro p-value ({g1_name})", f"Shapiro p-value ({g2_name})",
        "Levene p-value",
        f"Normality pass ({g1_name})", f"Normality pass ({g2_name})",
        "Variance equal pass",
        "Effect Size (Hedge's g)", "Large Effect",
        "Class_n", "Class_same_direction_n", "Class_same_direction_frac", "Class_median_log2FC",
        "Neighbor_n", "Neighbor_same_direction_n", "Neighbor_same_direction_frac", "Neighbor_median_log2FC",
        "Class_consistency_flag", "Neighbor_consistency_flag",
        "Significance",
    ]
    existing = [c for c in preferred_order if c in df.columns]
    remaining = [c for c in df.columns if c not in existing]
    df = df[existing + remaining]

    return df


def _export_all_test_pvalues(df: pd.DataFrame, out_path: Path) -> None:
    export_columns = [
        "UniqueID", "Annotation", "Annotation Type", "Headgroup", "Lipid Class",
        "Number of carbons in fatty acyls", "Double bond equivalents",
        "n_group1", "n_group2",
        "Auto selected test",
        "p-value (Auto)", "p-value (Student)", "p-value (Welch)", "p-value (Mann-Whitney)",
        "Mann-Whitney rank-biserial", "Mann-Whitney direction agrees", "Mann-Whitney robustness",
        "p-value",
    ]
    existing = [c for c in export_columns if c in df.columns]
    df.loc[:, existing].to_csv(out_path, index=False, encoding="utf-8-sig")

def _plot_volcano(
    df,
    g1,
    g2,
    save_dir: Path,
    p_thresh,
    fdr_thresh,
    fc_thresh,
    style: dict,
    sample_type: str = "",
    annotate_labels: bool = False,
):
    plt.close('all')
    df = df.copy()

    # Handle empty comparisons: still save a placeholder plot
    if df is None or len(df) == 0:
        fig, ax = plt.subplots(figsize=(10, 7), constrained_layout=False, facecolor="white")
        ax.set_facecolor("white")
        fig.subplots_adjust(left=0.12, right=0.98, top=0.86, bottom=0.20)

        ax.text(
            0.5, 0.5,
            f"No valid features to plot\n({g1} vs {g2})",
            ha="center", va="center", transform=ax.transAxes
        )
        ax.set_title(f"Volcano: {g1} vs {g2}", fontsize=style["title_size"], pad=15, fontweight="semibold")
        ax.set_xlabel("log2 fold change", fontsize=style["label_size"])
        ax.set_ylabel("-log10 FDR", fontsize=style["label_size"])
        ax.set_xlim(-1, 1)
        ax.set_ylim(0, 1)

        for spine in ax.spines.values():
            spine.set_visible(True)
            spine.set_linewidth(1.0)
            spine.set_color("black")

        fname_base = f"VolcanoFDR_{_sanitize_filename(g1)}_vs_{_sanitize_filename(g2)}_FC{_sanitize_filename(str(fc_thresh))}"
        _save_figure(fig, save_dir / f"{fname_base}.png", dpi=style["dpi"], facecolor=fig.get_facecolor())
        _save_figure(fig, save_dir / f"{fname_base}.svg", facecolor=fig.get_facecolor())
        plt.close(fig)
        return

    # Keep only finite rows
    df = df.replace([np.inf, -np.inf], np.nan).dropna(
        subset=["log2(Fold Change)", "-log10(FDR p-value)"]
    )

    # --- create a Figure and an Axes; draw everything on "ax" ---
    fig, ax = plt.subplots(figsize=(12, 8), constrained_layout=False, facecolor="white")
    ax.set_facecolor("white")
    # Reserve bottom space for the legend; do this ONCE.
    fig.subplots_adjust(left=0.12, right=0.76, top=0.86, bottom=0.30)

    log2fc_thresh = np.log2(fc_thresh)

    # Axis limits — use true min/max only (no percentiles)
    fc_line = np.log2(fc_thresh)

    # X limits
    x = df["log2(Fold Change)"].to_numpy()
    x_min = float(np.nanmin(x))
    x_max = float(np.nanmax(x))
    x_span = max(1e-6, x_max - x_min)
    x_pad  = max(0.5, 0.05 * x_span)
    ax.set_xlim(
        min(x_min - x_pad, -fc_line - 0.1),
        max(x_max + x_pad,  fc_line + 0.1),
    )

    # Y limits
    y = df["-log10(FDR p-value)"].to_numpy()
    y_max = float(np.nanmax(y)) if y.size else 0.0
    y_cut = -np.log10(fdr_thresh)
    y_pad = max(0.1, 0.08 * max(1.0, y_max))
    ax.set_ylim(0.0, max(y_max + y_pad, y_cut + 0.2, 1.25))

    df["Class_Color_Label"] = df.get("Lipid Class", pd.Series("Unknown", index=df.index)).map(_canon_class)
    class_palette = _build_class_palette(df["Class_Color_Label"].astype(str).tolist(), sample_type=sample_type)
    up = df[df["Significance"] == "Up"]
    down = df[df["Significance"] == "Down"]
    ns = df[df["Significance"] == "Not Significant"]

    # --- all scatters drawn on ax (no stray parens) ---
    ax.scatter(
        ns["log2(Fold Change)"],
        _add_jitter(ns["-log10(FDR p-value)"]),
        c=VOLCANO_NS,
        edgecolors="none", alpha=0.8, s=25, linewidth=0.0,
        label="Not Significant", zorder=1
    )
    if not down.empty:
        ax.scatter(
            down["log2(Fold Change)"],
            _add_jitter(down["-log10(FDR p-value)"]),
            c=[class_palette.get(cls, VOLCANO_DOWN) for cls in down["Class_Color_Label"].astype(str)],
            edgecolors="black", linewidth=0.35, alpha=0.95, s=style["marker_size"],
            marker="v", label="Down", zorder=3
        )
    if not up.empty:
        ax.scatter(
            up["log2(Fold Change)"],
            _add_jitter(up["-log10(FDR p-value)"]),
            c=[class_palette.get(cls, VOLCANO_UP) for cls in up["Class_Color_Label"].astype(str)],
            edgecolors="black", linewidth=0.35, alpha=0.95, s=style["marker_size"],
            marker="^", label="Up", zorder=4
        )


    # Threshold lines
    ax.axhline(y=-np.log10(fdr_thresh), color="gray", linestyle="--", linewidth=0.6, zorder=2)
    ax.axvline(x= log2fc_thresh,        color="gray", linestyle="--", linewidth=0.6)
    ax.axvline(x=-log2fc_thresh,        color="gray", linestyle="--", linewidth=0.6)

    # Labels / title
    ax.set_title(f"Volcano: {g1} vs {g2}", fontsize=style["title_size"], pad=15, fontweight="semibold")
    ax.set_xlabel("log2 fold change", fontsize=style["label_size"], labelpad=6)
    ax.set_ylabel("-log10 FDR", fontsize=style["label_size"], labelpad=6)

    ax.tick_params(labelsize=style["tick_size"])

    # Legend (bottom center)
    legend_labels = [
        f"Significantly Increased (FC {g1}/{g2} ≥ {fc_thresh:.2f}, p < {p_thresh}, FDR < {fdr_thresh}): {len(up)}",
        f"Significantly Decreased (FC {g1}/{g2} ≤ {1/fc_thresh:.2f}, p < {p_thresh}, FDR < {fdr_thresh}): {len(down)}",
        "Not Significant",
    ]
    handles = [
        Line2D([], [], marker='^', color='none', markerfacecolor="white", markeredgecolor="black", markersize=8, label=legend_labels[0]),
        Line2D([], [], marker='v', color='none', markerfacecolor="white", markeredgecolor="black", markersize=8, label=legend_labels[1]),
        Line2D([], [], marker='o', color='none', markerfacecolor=VOLCANO_NS, markeredgecolor=VOLCANO_NS, markersize=8, label=legend_labels[2]),
    ]
    # legend anchored below the axes but inside the reserved bottom margin
    fig.legend(
        handles=handles,
        loc="lower center",
        bbox_to_anchor=(0.5, 0.05),  # within the figure canvas
        ncol=1,
        frameon=True,
        fontsize=style["legend_size"]
    )

    sig_present_classes = _ordered_present_classes(
        df.loc[df["Significance"] != "Not Significant", "Class_Color_Label"].astype(str).tolist(),
        sample_type=sample_type,
    )
    if sig_present_classes:
        class_handles = [
            Patch(facecolor=class_palette[cls], edgecolor="black", linewidth=0.3, label=cls)
            for cls in sig_present_classes
        ]
        fig.legend(
            handles=class_handles,
            loc="upper left",
            bbox_to_anchor=(0.77, 0.91),  # (x, y)
            ncol=1,
            frameon=False,
            fontsize=max(style["legend_size"] - 2, 8),
        )

    label_df = df.copy()
    if annotate_labels and not label_df.empty:
        label_df["abs_log2FC"] = pd.to_numeric(label_df["log2(Fold Change)"], errors="coerce").abs()
        label_df["abs_effect_size"] = pd.to_numeric(
            label_df.get("Effect Size (Hedge's g)", pd.Series(np.nan, index=label_df.index)),
            errors="coerce",
        ).abs().fillna(0.0)
        label_df["priority_raw_p"] = pd.to_numeric(label_df.get("p-value"), errors="coerce").lt(0.05)
        label_df["priority_large_effect"] = label_df.get("Large Effect", pd.Series("No", index=label_df.index)).astype(str).eq("Yes")
        label_df["priority_coherence"] = (
            label_df.get("Class_consistency_flag", pd.Series("No", index=label_df.index)).astype(str).eq("Yes")
            | label_df.get("Neighbor_consistency_flag", pd.Series("No", index=label_df.index)).astype(str).eq("Yes")
        )
        priority_mask = (
            label_df["priority_raw_p"]
            | label_df["priority_large_effect"]
            | label_df["priority_coherence"]
        )
        if priority_mask.any():
            label_df = label_df.loc[priority_mask].copy()
        label_df = label_df.sort_values(
            by=[
                "priority_coherence",
                "priority_large_effect",
                "priority_raw_p",
                "abs_effect_size",
                "abs_log2FC",
                "p-value",
            ],
            ascending=[False, False, False, False, False, True],
        ).head(8)
        for _, row in label_df.iterrows():
            label = str(row.get("Annotation") or row.get("Headgroup") or row.get("UniqueID"))
            ax.text(
                float(row["log2(Fold Change)"]),
                float(row["-log10(FDR p-value)"]) + 0.04,
                label,
                fontsize=max(style["tick_size"] - 1, 8),
                ha="left",
                va="bottom",
                color="black",
            )

    # Add full rectangular border
    for spine in ax.spines.values():
        spine.set_visible(True)
        spine.set_linewidth(style["line_width"])
        spine.set_color("black")

    # Save
    fname_base = f"VolcanoFDR_{_sanitize_filename(g1)}_vs_{_sanitize_filename(g2)}_FC{_sanitize_filename(str(fc_thresh))}"
    _save_figure(fig, save_dir / f"{fname_base}.png", dpi=style["dpi"], facecolor=fig.get_facecolor())
    _save_figure(fig, save_dir / f"{fname_base}.svg", facecolor=fig.get_facecolor())
    
    plt.close(fig)
    
# ==========================================================
# Summary tables (CSV + Excel)
# ==========================================================
def _save_summary_tables(csv_dir: Path, output_csv: Path, output_excel: Path,
                         fc_thresh: float, fdr_thresh: float, p_thresh: float):
    from openpyxl.utils import get_column_letter

    files = sorted([f for f in os.listdir(csv_dir) if f.endswith("_FDR.csv")])
    if not files:
        print("[Volcano] No _FDR.csv files found.", flush=True)
        return

    meta_cols = ["UniqueID", "Annotation", "Annotation Type", "Headgroup", "Lipid Class"]
    base_numeric_cols = [
        "Fold Change",
        "log2(Fold Change)",
        "p-value",
        "FDR p-value",
        "p-value (Mann-Whitney)",
        "Mann-Whitney rank-biserial",
        "Mann-Whitney robustness",
        "Levene p-value",
        "Effect Size (Hedge's g)",
        "Large Effect",
    ]
    comparison_order = [
        "FoldChange",
        "log2FC",
        "pval",
        "FDR_p",
        "Sig_FDR_and_p",
        "Sig_p_only",
        "MannWhitney_p",
        "MannWhitney_rbc",
        "Nonparametric_support",
        "Levene_p",
    ]
    header_label_map = {
        "FoldChange": "Fold Change",
        "log2FC": "log₂(Fold Change)",
        "pval": "raw p-value",
        "FDR_p": "FDR-p",
        "Sig_FDR_and_p": "Significant (FDR-p & raw p)",
        "Sig_p_only": "Significant (raw p only)",
        "MannWhitney_p": "Mann-Whitney p-value",
        "MannWhitney_rbc": "Mann-Whitney rank-biserial",
        "Nonparametric_support": "Nonparametric support",
        "Levene_p": "Levene p-value",
        "EffectSize": "Effect Size (Hedge's g)",
        "LargeEffect": "Large Effect",
    }

    combined_meta = None
    combined_numeric = None
    summaries = []

    for fname in files:
        comp = fname.replace("_FDR.csv", "")
        df = pd.read_csv(csv_dir / fname)

        if "UniqueID" not in df.columns:
            print(f"[Volcano] Skipping {fname}: missing UniqueID.", flush=True)
            continue

        df["UniqueID"] = df["UniqueID"].astype(str).str.strip()

        for c in meta_cols:
            if c not in df.columns:
                df[c] = ""

        for c in base_numeric_cols:
            if c not in df.columns:
                df[c] = np.nan

        shapiro_cols = [c for c in df.columns if str(c).startswith("Shapiro p-value (")]
        normality_cols = [c for c in df.columns if str(c).startswith("Normality pass (")]

        assumption_cols = []
        for shapiro_col in shapiro_cols:
            suffix = shapiro_col[len("Shapiro p-value "):]
            paired_normality = f"Normality pass {suffix}"
            assumption_cols.append(shapiro_col)
            if paired_normality in normality_cols:
                assumption_cols.append(paired_normality)
        for normality_col in normality_cols:
            if normality_col not in assumption_cols:
                assumption_cols.append(normality_col)

        up = (df["Fold Change"] > fc_thresh) & (df["FDR p-value"] < fdr_thresh) & (df["p-value"] < p_thresh)
        down = (df["Fold Change"] < 1 / fc_thresh) & (df["FDR p-value"] < fdr_thresh) & (df["p-value"] < p_thresh)
        summaries.append((comp, int(np.nansum(up)), int(np.nansum(down))))

        this_meta = df[meta_cols].drop_duplicates("UniqueID").set_index("UniqueID")
        combined_meta = this_meta if combined_meta is None else combined_meta.combine_first(this_meta)

        use = df[["UniqueID"] + base_numeric_cols + assumption_cols].copy()
        use.rename(columns={
            "Fold Change": "FoldChange",
            "log2(Fold Change)": "log2FC",
            "p-value": "pval",
            "FDR p-value": "FDR_p",
            "p-value (Mann-Whitney)": "MannWhitney_p",
            "Mann-Whitney rank-biserial": "MannWhitney_rbc",
            "Mann-Whitney robustness": "Nonparametric_support",
            "Levene p-value": "Levene_p",
            "Effect Size (Hedge's g)": "EffectSize",
            "Large Effect": "LargeEffect",
        }, inplace=True)

        for c in ["FoldChange", "log2FC", "pval", "FDR_p", "MannWhitney_p", "MannWhitney_rbc", "Levene_p", "EffectSize"] + shapiro_cols:
            if c in use.columns:
                use[c] = pd.to_numeric(use[c], errors="coerce")

        use["Sig_FDR_and_p"] = np.select(
            [
                (use["FDR_p"] < fdr_thresh) & (use["pval"] < p_thresh) & (use["FoldChange"] > 1),
                (use["FDR_p"] < fdr_thresh) & (use["pval"] < p_thresh) & (use["FoldChange"] < 1),
            ],
            ["Significantly increased", "Significantly decreased"],
            default="Not significant",
        )
        use["Sig_p_only"] = np.select(
            [
                (use["pval"] < p_thresh) & (use["FoldChange"] > 1),
                (use["pval"] < p_thresh) & (use["FoldChange"] < 1),
            ],
            ["Significantly increased", "Significantly decreased"],
            default="Not significant",
        )

        ordered_cols = (
            [c for c in comparison_order if c in use.columns]
            + [c for c in assumption_cols if c in use.columns]
            + [c for c in ["EffectSize", "LargeEffect"] if c in use.columns]
        )
        use = use[["UniqueID"] + ordered_cols].set_index("UniqueID")
        use.columns = pd.MultiIndex.from_product([[comp], list(use.columns)])
        combined_numeric = use if combined_numeric is None else combined_numeric.join(use, how="outer")

    if combined_meta is None:
        print("[Volcano] No valid files for summary.", flush=True)
        return

    if combined_meta.index.name != "UniqueID":
        combined_meta.index.name = "UniqueID"

    combined_wide = combined_meta.copy() if combined_numeric is None else pd.concat([combined_meta, combined_numeric], axis=1, join="outer")

    combined_csv = combined_wide.reset_index()

    def _flatten_col(c):
        if isinstance(c, tuple):
            return f"{c[0]}__{c[1]}"
        return str(c)

    combined_csv.columns = [_flatten_col(c) for c in combined_csv.columns]
    combined_csv.to_csv(output_csv, index=False, encoding="utf-8-sig")

    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Volcano Data"
    comparison_blocks = []

    meta_headers_row1 = [""] * len(meta_cols)
    meta_headers_row2 = meta_cols

    if combined_numeric is None:
        ws1.append(meta_headers_row1)
        ws1.append(meta_headers_row2)
        for _, r in combined_meta.reset_index()[meta_cols].iterrows():
            ws1.append(r.tolist())
    else:
        numeric_multi = combined_numeric.columns
        row1 = meta_headers_row1 + [c[0] for c in numeric_multi]
        row2 = meta_headers_row2 + [header_label_map.get(c[1], c[1]) for c in numeric_multi]
        ws1.append(row1)
        ws1.append(row2)

        full = pd.concat([combined_meta.copy(), combined_numeric.copy()], axis=1, join="outer").reset_index()
        for _, r in full.iterrows():
            ws1.append(r.tolist())

        start_col = len(meta_cols) + 1
        block_start = start_col
        last_comp = row1[start_col - 1] if len(row1) >= start_col else None
        for col_idx in range(start_col, start_col + len(numeric_multi)):
            comp_name = row1[col_idx - 1]
            if comp_name != last_comp:
                comparison_blocks.append((block_start, col_idx - 1))
                ws1.merge_cells(start_row=1, start_column=block_start, end_row=1, end_column=col_idx - 1)
                block_start = col_idx
                last_comp = comp_name
        if last_comp is not None:
            comparison_blocks.append((block_start, start_col + len(numeric_multi) - 1))
            ws1.merge_cells(
                start_row=1,
                start_column=block_start,
                end_row=1,
                end_column=start_col + len(numeric_multi) - 1,
            )

    header_font = Font(bold=True)
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row_idx in (1, 2):
        for cell in ws1[row_idx]:
            cell.font = header_font
            cell.alignment = center_align

    comparison_fills = [
        "FCE4D6",
        "E2F0D9",
        "DDEBF7",
        "FFF2CC",
        "E4DFEC",
        "D9EAF7",
        "F4CCCC",
        "D9F0E3",
        "F9E2AF",
        "D6EAF8",
    ]
    thin_side = Side(style="thin", color="BFBFBF")
    medium_side = Side(style="medium", color="7F7F7F")
    max_row = ws1.max_row
    max_col = ws1.max_column

    for col_idx in range(1, len(meta_cols) + 1):
        for row_idx in range(1, max_row + 1):
            cell = ws1.cell(row=row_idx, column=col_idx)
            cell.border = Border(
                left=medium_side if col_idx == 1 else thin_side,
                right=medium_side if col_idx == len(meta_cols) else thin_side,
                top=medium_side if row_idx in (1, 2) else thin_side,
                bottom=medium_side if row_idx == max_row else thin_side,
            )

    if combined_numeric is not None:
        fill_idx = 0
        for block_start, block_end in comparison_blocks:
            fill = PatternFill(fill_type="solid", fgColor=comparison_fills[fill_idx % len(comparison_fills)])
            for col_idx in range(block_start, block_end + 1):
                for row_idx in range(1, max_row + 1):
                    cell = ws1.cell(row=row_idx, column=col_idx)
                    cell.fill = fill
                    cell.border = Border(
                        left=medium_side if col_idx == block_start else thin_side,
                        right=medium_side if col_idx == block_end else thin_side,
                        top=medium_side if row_idx in (1, 2) else thin_side,
                        bottom=medium_side if row_idx == max_row else thin_side,
                    )
            fill_idx += 1

    for col_idx, column_cells in enumerate(ws1.columns, start=1):
        max_length = 0
        for cell in column_cells:
            if cell.value is None:
                continue
            max_length = max(max_length, len(str(cell.value)))
        ws1.column_dimensions[get_column_letter(col_idx)].width = min(max_length + 2, 40)

    ws2 = wb.create_sheet("Summary Counts")
    ws2.append([
        "Comparison",
        f"Up (FC ≥ {fc_thresh}, raw p < {p_thresh}, FDR < {fdr_thresh})",
        f"Down (FC ≤ 1/{fc_thresh}, raw p < {p_thresh}, FDR < {fdr_thresh})",
    ])
    for comp, up_cnt, down_cnt in summaries:
        ws2.append([comp, up_cnt, down_cnt])

    for col_idx, column_cells in enumerate(ws2.columns, start=1):
        max_length = 0
        for cell in column_cells:
            if cell.value is None:
                continue
            max_length = max(max_length, len(str(cell.value)))
        ws2.column_dimensions[get_column_letter(col_idx)].width = min(max_length + 2, 40)

    wb.save(output_excel)
    print(f"[Volcano] Saved Excel with two sheets: {output_excel}", flush=True)

# ==========================================================
# Class extraction / grouping (for bar + bubble plots)
# ==========================================================
def _extract_Class(name):
    """Extract lipid Class token from your 'UniqueID' naming."""
    if pd.isna(name):
        return "Unknown"
    name = str(name)
    if "|" in name:
        lipid_part = name.split("|")[-1].strip()
    else:
        lipid_part = name.strip()
    m = re.match(r"^([A-Z]+(?: O)?(?:-)?[A-Z]*)", lipid_part)
    return m.group(1) if m else "Unknown"

# Master class order
_CLASS_ORDER_DEFAULT = _CLASS_ORDER

def _resolve_class_order(sample_type: str) -> list[str]:
    st = (sample_type or "").strip().lower()
    if st.startswith("bact"):
        return _CLASS_ORDER_BACTERIA
    elif st.startswith("mamm"):
        return _CLASS_ORDER_MAMMALIAN
    return _CLASS_ORDER_DEFAULT


def _canon_class(x: str) -> str:
    x = str(x or "").strip()
    if x in _CLASS_ORDER or x in _CLASS_ORDER_BACTERIA or x in _CLASS_ORDER_MAMMALIAN:
        return x
    if x in _CLASS_GROUP_MAP:
        return _CLASS_GROUP_MAP[x]
    return x if x else "Unknown"


def _build_class_palette(classes: list[str], sample_type: Optional[str] = None) -> dict[str, tuple[float, float, float]]:
    present_classes = []
    for cls in classes:
        cls = _canon_class(cls)
        if cls and cls not in present_classes:
            present_classes.append(cls)

    if not present_classes:
        return {}

    preferred_order = []
    for cls in _resolve_class_order(sample_type or ""):
        canon = _canon_class(cls)
        if canon and canon not in preferred_order:
            preferred_order.append(canon)

    # Assign colors from the full canonical class order first so a class keeps
    # the same color even when a plot only contains a subset of classes.
    extra_classes = sorted([cls for cls in present_classes if cls not in preferred_order])
    full_color_order = preferred_order + extra_classes

    # Curated high-contrast palette so neighboring classes remain visually distinct
    # while staying fixed across volcano and bubble plots.
    wheel = [
        "#E15759", "#4E79A7", "#F28E2B", "#76B7B2", "#59A14F",
        "#EDC948", "#B07AA1", "#FF9DA7", "#9C755F", "#BAB0AC",
        "#1F77B4", "#FF7F0E", "#2CA02C", "#D62728", "#9467BD",
        "#8C564B", "#E377C2", "#7F7F7F", "#BCBD22", "#17BECF",
        "#AEC7E8", "#FFBB78", "#98DF8A", "#FF9896", "#C5B0D5",
        "#C49C94", "#F7B6D2", "#C7C7C7", "#DBDB8D", "#9EDAE5",
    ]
    full_palette = {cls: mpl.colors.to_rgb(wheel[i % len(wheel)]) for i, cls in enumerate(full_color_order)}
    return {cls: full_palette[cls] for cls in present_classes}


def _ordered_present_classes(classes: list[str], sample_type: Optional[str] = None) -> list[str]:
    present_classes = []
    for cls in classes:
        cls = _canon_class(cls)
        if cls and cls not in present_classes:
            present_classes.append(cls)

    preferred_order = [_canon_class(cls) for cls in _resolve_class_order(sample_type or "")]
    ordered_classes = [cls for cls in preferred_order if cls in present_classes]
    ordered_classes.extend([cls for cls in present_classes if cls not in ordered_classes])
    return ordered_classes


# ==========================================================
# Complementary plots
# ==========================================================
def _generate_ranked_effect_plots_from_csv(input_dir: Path, output_dir: Path, sample_type: str):
    print("\n[Volcano] Generating ranked effect plots", flush=True)
    plt.close('all')
    csv_input = input_dir / "CSV_files"
    if not csv_input.exists():
        print(f"[Volcano] Skipping ranked effect plots: missing folder {csv_input}", flush=True)
        return

    _ensure_dir(output_dir)
    files = [f for f in os.listdir(csv_input) if f.endswith("_FDR.csv")]

    for file in files:
        df = pd.read_csv(csv_input / file)
        if df.empty or "UniqueID" not in df.columns:
            continue

        if "Lipid Class" in df.columns:
            df["Class_Group"] = (
                df["Lipid Class"].astype(str).str.strip().replace("", "Unknown").replace("nan", "Unknown")
            )
        else:
            df["Class_Group"] = df["UniqueID"].apply(_extract_Class)
        df["Class_Group"] = df["Class_Group"].map(_canon_class)

        effect_vals = pd.to_numeric(
            df.get("Effect Size (Hedge's g)", pd.Series(np.nan, index=df.index)),
            errors="coerce",
        )
        if effect_vals.notna().sum() >= 3:
            y_col = "Effect Size (Hedge's g)"
            rank_basis = effect_vals.abs()
        else:
            y_col = "log2(Fold Change)"
            rank_basis = pd.to_numeric(df.get("log2(Fold Change)"), errors="coerce").abs()

        plot_df = df.copy()
        plot_df[y_col] = pd.to_numeric(plot_df.get(y_col), errors="coerce")
        plot_df["_rank_basis"] = pd.to_numeric(rank_basis, errors="coerce")
        plot_df = plot_df.dropna(subset=[y_col, "_rank_basis", "Class_Group"])
        if plot_df.empty:
            continue

        plot_df = plot_df.sort_values("_rank_basis", ascending=False).reset_index(drop=True)
        plot_df["Rank"] = np.arange(1, len(plot_df) + 1)
        top_n = max(1, int(np.ceil(0.10 * len(plot_df))))
        plot_df["TopDecile"] = False
        plot_df.loc[:top_n - 1, "TopDecile"] = True

        class_palette = _build_class_palette(plot_df["Class_Group"].astype(str).tolist(), sample_type=sample_type)
        colors = [class_palette.get(cls, (0.5, 0.5, 0.5)) for cls in plot_df["Class_Group"].astype(str)]

        fig, ax = plt.subplots(figsize=(13, 7.5), facecolor="white")
        ax.set_facecolor("white")

        ax.scatter(
            plot_df.loc[~plot_df["TopDecile"], "Rank"],
            plot_df.loc[~plot_df["TopDecile"], y_col],
            c=[colors[i] for i in np.where(~plot_df["TopDecile"])[0]],
            s=34,
            alpha=0.82,
            linewidth=0.35,
            edgecolor="white",
            zorder=2,
        )
        ax.scatter(
            plot_df.loc[plot_df["TopDecile"], "Rank"],
            plot_df.loc[plot_df["TopDecile"], y_col],
            c=[colors[i] for i in np.where(plot_df["TopDecile"])[0]],
            s=62,
            alpha=0.96,
            linewidth=0.9,
            edgecolor="black",
            zorder=3,
            label="Top 10% by |effect|",
        )

        label_candidates = plot_df.loc[plot_df["TopDecile"]].copy()
        label_candidates = pd.concat([
            label_candidates.loc[label_candidates[y_col] > 0].head(5),
            label_candidates.loc[label_candidates[y_col] < 0].head(5),
        ], axis=0).drop_duplicates().copy()
        if not label_candidates.empty:
            y_all = pd.to_numeric(plot_df[y_col], errors="coerce")
            y_min = float(y_all.min())
            y_max = float(y_all.max())
            y_span = max(1e-6, y_max - y_min)
            min_sep = max(0.18, 0.025 * y_span)
            max_shift = max(0.35, 0.08 * y_span)
            x_pad = max(1.8, 0.010 * len(plot_df))

            def _spread_positions(group_df: pd.DataFrame, positive: bool) -> list[tuple[int, float]]:
                if group_df.empty:
                    return []
                group_df = group_df.sort_values(y_col, ascending=False).copy()
                placed = []
                used_y = []
                for idx, row in group_df.iterrows():
                    y_target = float(row[y_col])
                    candidates = [y_target]
                    for step in range(1, 8):
                        delta = step * min_sep
                        candidates.extend([y_target + delta, y_target - delta])
                    lower = y_target - max_shift
                    upper = y_target + max_shift
                    y_text = y_target
                    for cand in candidates:
                        if cand < lower or cand > upper:
                            continue
                        if all(abs(cand - prev) >= min_sep for prev in used_y):
                            y_text = cand
                            break
                    placed.append((idx, y_text))
                    used_y.append(y_text)
                return placed

            pos_positions = _spread_positions(label_candidates[label_candidates[y_col] >= 0], positive=True)
            neg_positions = _spread_positions(label_candidates[label_candidates[y_col] < 0], positive=False)

            for idx, y_text in pos_positions + neg_positions:
                row = label_candidates.loc[idx]
                y_val = float(row[y_col])
                label = str(row.get("Annotation") or row.get("Headgroup") or row.get("UniqueID"))
                use_arrow = abs(y_text - y_val) > (0.35 * min_sep)
                ax.annotate(
                    label,
                    xy=(float(row["Rank"]), y_val),
                    xytext=(float(row["Rank"]) + x_pad, y_text),
                    textcoords="data",
                    fontsize=8.5,
                    ha="left",
                    va="center",
                    color="#1F1F1F",
                    arrowprops=(
                        dict(
                            arrowstyle="-",
                            color="#777777",
                            lw=0.6,
                            alpha=0.75,
                            shrinkA=0,
                            shrinkB=4,
                        ) if use_arrow else None
                    ),
                    bbox=dict(boxstyle="round,pad=0.15", facecolor="white", edgecolor="none", alpha=0.78),
                    zorder=4,
                )

        ax.axhline(0, color="#666666", linestyle="--", linewidth=0.8, zorder=1)
        ax.grid(axis="y", color="#D9D9D9", linewidth=0.7, alpha=0.8)
        ax.set_axisbelow(True)

        comparison = file.replace("_FDR.csv", "")
        y_label = "Effect Size (Hedge's g)" if y_col == "Effect Size (Hedge's g)" else "log2(Fold Change)"
        basis_label = "|effect size|" if y_col == "Effect Size (Hedge's g)" else "|log2FC|"
        ax.set_title(f"Ranked {y_label} profile\n({comparison})", fontsize=20, pad=16)
        ax.set_xlabel(f"Feature rank by {basis_label}", fontsize=16, labelpad=10)
        ax.set_ylabel(y_label, fontsize=16, labelpad=10)
        ax.tick_params(axis="both", labelsize=12)

        handles = [
            Patch(facecolor=class_palette[cls], edgecolor="black", linewidth=0.3, label=cls)
            for cls in _ordered_present_classes(plot_df["Class_Group"].astype(str).tolist(), sample_type=sample_type)
        ]
        if handles:
            class_legend = ax.legend(
                handles=handles,
                title="Lipid class",
                fontsize=10,
                title_fontsize=11,
                loc="upper left",
                bbox_to_anchor=(1.01, 1.0),
                frameon=False,
            )
            ax.add_artist(class_legend)
        ax.legend(loc="upper left", bbox_to_anchor=(1.01, 0.20), frameon=False, fontsize=10)

        for spine in ax.spines.values():
            spine.set_visible(True)
            spine.set_linewidth(1.0)
            spine.set_color("black")

        fig.tight_layout()
        out_png = output_dir / file.replace("_FDR.csv", "_ranked_effect_plot.png")
        out_svg = output_dir / file.replace("_FDR.csv", "_ranked_effect_plot.svg")
        _save_figure(fig, out_png, dpi=140, bbox_inches="tight", facecolor=fig.get_facecolor())
        _save_figure(fig, out_svg, bbox_inches="tight", facecolor=fig.get_facecolor())
        plt.close(fig)


def _generate_class_directionality_plots_from_csv(input_dir: Path, output_dir: Path, sample_type: str):
    print("\n[Volcano] Generating class directionality plots", flush=True)
    plt.close('all')
    csv_input = input_dir / "CSV_files"
    if not csv_input.exists():
        print(f"[Volcano] Skipping class directionality plots: missing folder {csv_input}", flush=True)
        return

    _ensure_dir(output_dir)
    files = [f for f in os.listdir(csv_input) if f.endswith("_FDR.csv")]

    for file in files:
        df = pd.read_csv(csv_input / file)
        if df.empty or "UniqueID" not in df.columns or "log2(Fold Change)" not in df.columns:
            continue

        if "Lipid Class" in df.columns:
            df["Class_Group"] = (
                df["Lipid Class"].astype(str).str.strip().replace("", "Unknown").replace("nan", "Unknown")
            )
        else:
            df["Class_Group"] = df["UniqueID"].apply(_extract_Class)
        df["Class_Group"] = df["Class_Group"].map(_canon_class)
        df["log2(Fold Change)"] = pd.to_numeric(df["log2(Fold Change)"], errors="coerce")
        df["p-value"] = pd.to_numeric(df.get("p-value"), errors="coerce")

        plot_df = df.dropna(subset=["Class_Group", "log2(Fold Change)"]).copy()
        if plot_df.empty:
            continue

        summary = (
            plot_df.groupby("Class_Group", dropna=False)
            .apply(
                lambda g: pd.Series({
                    "Total_n": int(len(g)),
                    "Increased_n": int(g["log2(Fold Change)"].gt(0).sum()),
                    "Decreased_n": int(g["log2(Fold Change)"].lt(0).sum()),
                    "Significant_n": int(g["p-value"].lt(0.05).sum()),
                    "Same_direction_frac": float(max(g["log2(Fold Change)"].gt(0).sum(), g["log2(Fold Change)"].lt(0).sum()) / len(g)) if len(g) else np.nan,
                })
            )
            .reset_index()
        )
        summary = summary.loc[summary["Total_n"] > 0].copy()
        if summary.empty:
            continue

        summary["Increased_frac"] = summary["Increased_n"] / summary["Total_n"]
        summary["Decreased_frac"] = summary["Decreased_n"] / summary["Total_n"]
        summary["Significant_frac"] = summary["Significant_n"] / summary["Total_n"]

        class_order = _ordered_present_classes(summary["Class_Group"].astype(str).tolist(), sample_type=sample_type)
        summary["Class_Group"] = pd.Categorical(summary["Class_Group"], categories=class_order, ordered=True)
        summary = summary.sort_values("Class_Group").reset_index(drop=True)

        x = np.arange(len(summary))
        fig, ax = plt.subplots(figsize=(max(12, 0.55 * len(summary) + 4), 7.8), facecolor="white")
        ax.set_facecolor("white")

        up_color = "#D1495B"
        down_color = "#2F6690"
        ax.bar(x, summary["Increased_frac"], color=up_color, width=0.82, label="Increased fraction", zorder=2)
        ax.bar(
            x,
            summary["Decreased_frac"],
            bottom=summary["Increased_frac"],
            color=down_color,
            width=0.82,
            label="Decreased fraction",
            zorder=2,
        )

        for i, xpos in enumerate(x):
            if i % 2 == 0:
                ax.axvspan(xpos - 0.5, xpos + 0.5, color="#F3F3F3", alpha=0.8, zorder=0)

        ax2 = ax.twinx()
        ax2.plot(
            x,
            summary["Significant_frac"],
            color="#222222",
            linewidth=1.4,
            marker="o",
            markersize=5.5,
            label="raw p < 0.05 fraction",
            zorder=4,
        )
        ax2.vlines(x, 0, summary["Significant_frac"], color="#222222", linewidth=1.0, alpha=0.75, zorder=3)

        for xpos, frac in zip(x, summary["Same_direction_frac"]):
            if np.isfinite(frac):
                ax.text(
                    xpos,
                    1.015,
                    f"{frac:.2f}",
                    ha="center",
                    va="bottom",
                    fontsize=9,
                    color="#444444",
                    rotation=90 if len(summary) > 12 else 0,
                )

        comparison = file.replace("_FDR.csv", "")
        ax.set_title(f"Class directionality profile\n({comparison})", fontsize=20, pad=16)
        ax.set_ylabel("Fraction of detected lipids", fontsize=15, labelpad=10)
        ax2.set_ylabel("Fraction of detected lipids with raw p < 0.05", fontsize=15, labelpad=12)
        ax.set_xlabel("Lipid class", fontsize=15, labelpad=10)
        ax.set_xticks(x)
        ax.set_xticklabels(summary["Class_Group"].astype(str).tolist(), rotation=90, fontsize=11)
        ax.set_ylim(0, 1.08)
        ax2.set_ylim(0, max(1.02, float(summary["Significant_frac"].max()) * 1.12 if len(summary) else 1.02))
        ax.tick_params(axis="y", labelsize=12)
        ax2.tick_params(axis="y", labelsize=12)
        ax.grid(axis="y", color="#D9D9D9", linewidth=0.7, alpha=0.8)
        ax.set_axisbelow(True)
        ax.text(
            0.995,
            1.045,
            "Same direction fraction",
            transform=ax.transAxes,
            ha="right",
            va="bottom",
            fontsize=10,
            color="#444444",
        )

        handles1, labels1 = ax.get_legend_handles_labels()
        handles2, labels2 = ax2.get_legend_handles_labels()
        ax.legend(
            handles1 + handles2,
            labels1 + labels2,
            loc="upper left",
            bbox_to_anchor=(1.08, 1.0),
            frameon=False,
            fontsize=10,
        )

        for spine in ax.spines.values():
            spine.set_visible(True)
            spine.set_linewidth(1.0)
            spine.set_color("black")
        for spine in ax2.spines.values():
            spine.set_linewidth(1.0)
            spine.set_color("black")

        fig.tight_layout()
        out_png = output_dir / file.replace("_FDR.csv", "_class_directionality_plot.png")
        out_svg = output_dir / file.replace("_FDR.csv", "_class_directionality_plot.svg")
        _save_figure(fig, out_png, dpi=140, bbox_inches="tight", facecolor=fig.get_facecolor())
        _save_figure(fig, out_svg, bbox_inches="tight", facecolor=fig.get_facecolor())
        plt.close(fig)


# ==========================================================
# Class bar plots (Up black / Down gray)
# ==========================================================
def _generate_class_barplots_from_csv(input_dir: Path, output_dir: Path, sample_type: str,
                                      p_value_threshold: float, fdr_threshold: float, fold_change_threshold: float):
    print("\n[Volcano] Generating class bar plots", flush = True)
    plt.close('all')
    csv_input = input_dir / "CSV_files"
    if not csv_input.exists():
        print(f"[Volcano] Skipping plots: missing folder {csv_input}", flush=True)
        return
    files = [p.name for p in csv_input.glob("*_FDR.csv")]

    # class order by sample_type
    class_order = _resolve_class_order(sample_type)

    _ensure_dir(output_dir)

    for file in files:
        df = pd.read_csv(csv_input / file)

        # --- Robust checks and normalization ---
        if "UniqueID" not in df.columns or "Significance" not in df.columns:
            print(f"[Volcano] Skipping {file} (missing columns)", flush = True)
            continue

        # Normalize 'Significance' labels (avoid trailing spaces or lowercase)
        df["Significance"] = df["Significance"].astype(str).str.strip().str.capitalize()

        if "Significance" in df.columns:
            df["Significance"] = (
                df["Significance"]
                .astype(str)
                .str.strip()          # remove leading/trailing spaces
                .str.replace(r"\s+", " ", regex=True)  # collapse weird whitespace
                .str.capitalize()     # unify case ("up"→"Up", "down"→"Down")
            )
    
        # Ensure expected values exist
        valid_sig = df["Significance"].isin(["Up", "Down"])
        print(f"[Volcano] {file}: {valid_sig.sum()} significant entries detected", flush = True)
        if valid_sig.sum() == 0:
            print(f"[Volcano] {file}: 0 significant entries; producing empty plot", flush=True)
            # Create an empty but labeled figure so folders are consistent
            fig, ax = plt.subplots(figsize=(10, 4), facecolor="white")
            ax.set_facecolor("white")
            ax.text(0.5, 0.5, "No significant features\n(at current thresholds)",
                    ha="center", va="center", transform=ax.transAxes)
            ax.axis("off")
            outname = output_dir / file.replace("_FDR.csv", "_EMPTY.png")
            _save_figure(fig, outname, dpi=120, facecolor=fig.get_facecolor())
            plt.close(fig)
            # and continue to next file
            continue

        # --- Prefer 'Lipid Class' column from metadata ---
        if "Lipid Class" in df.columns:
            df["Class_Group"] = (
                df["Lipid Class"]
                .astype(str)
                .str.strip()
                .replace("", "Unknown")
                .replace("nan", "Unknown")
            )
        else:
            # fallback to parsed class from UniqueID
            df["Class"] = df["UniqueID"].apply(_extract_Class)
            df["Class_Group"] = df["Class"].map(_CLASS_GROUP_MAP).fillna(df["Class"])
        # -------------------------------------------------

        # --- Stable counts (no NaNs, fixed order) ---
        # --- Stable counts even if one column is missing ---
        count_data = (
            df.groupby(["Class_Group", "Significance"])
            .size()
            .unstack(fill_value=0)
        )
        # Ensure both columns exist
        for col in ("Up", "Down"):
            if col not in count_data.columns:
                count_data[col] = 0

        # Order columns and classes
        count_data = count_data[["Up", "Down"]].reindex(class_order, fill_value=0)

        # Up positive, Down negative for diverging bars
        count_data["Up"] = count_data["Up"].astype(int)
        count_data["Down"] = -count_data["Down"].astype(int)

        # --- Plot with robust limits ---
        # --- Figure/axes with explicit white backgrounds ---
        fig, ax = plt.subplots(figsize=(12, 6), facecolor="white")
        ax.set_facecolor("white")

        x = np.arange(len(count_data))
        ax.bar(x, count_data["Up"].to_numpy(), color="black", width=0.8)
        ax.bar(x, count_data["Down"].to_numpy(), color="gray", width=0.8)
        ax.set_xticks(x)
        ax.set_xticklabels(count_data.index)
        # anchor and pad so 90° labels don’t get clipped
        labels = ax.get_xticklabels()
        plt.setp(labels, rotation=90, ha="right", rotation_mode="anchor")
        ax.tick_params(axis="x", pad=8)
        
        # Alternating background bands (strongest readability gain)
        for i, xpos in enumerate(x):
            if i % 2 == 0:
                ax.axvspan(xpos - 0.5, xpos + 0.5, alpha=0.1, zorder=0,  color="lightgray")  
        ax.set_axisbelow(True)

        comparison = file.replace("_FDR.csv", "")
        ax.set_title(f"Class distribution for significantly altered lipids ({comparison})", pad=28)

        black_label = rf'raw p < {p_value_threshold}, FDR < {fdr_threshold}, FC ≥ {fold_change_threshold:.2f}'
        gray_label  = rf'raw p < {p_value_threshold}, FDR < {fdr_threshold}, FC ≤ {1/fold_change_threshold:.2f}'

        handles = [
            plt.matplotlib.patches.Patch(color="black", label=black_label),
            plt.matplotlib.patches.Patch(color="gray",  label=gray_label),
        ]
        fig.legend(
            handles=handles,
            loc="upper center",
            bbox_to_anchor=(0.5, 0.90),          # relative to the *figure*
            bbox_transform=fig.transFigure,
            ncol=2,
            frameon=False,
            fontsize=10
        )

        ax.axhline(0, color="black", linewidth=0.8)

        # Robust y-limits
        abs_counts = np.abs(count_data[["Up", "Down"]].to_numpy().ravel())
        ymax = int(np.nanpercentile(abs_counts, 99)) if abs_counts.size else 0
        ymax = max(ymax, int(abs_counts.max()) if abs_counts.size else 0, 1)
        ax.set_ylim(-ymax - 1, ymax + 1)

        ax.set_ylabel("Number of Significantly Altered Lipids")

        # Keep room for tick labels and legend; save on white background (no transparency)
        fig.subplots_adjust(bottom=0.28, top=0.84)

        # Add full rectangular border
        for spine in ax.spines.values():
            spine.set_visible(True)
            spine.set_linewidth(1.0)
            spine.set_color("black")

        outpath = output_dir / f"{comparison}_class_distribution_barplot.png"
        _save_figure(fig, outpath, dpi=120, facecolor=fig.get_facecolor(), bbox_inches="tight", pad_inches=0.25)
        plt.close(fig)


# ==========================================================
# Bubble plot (design 1): y = Class, color by log2FC (seismic)
# ==========================================================
def _generate_bubble_plot_from_csv(input_dir: Path, output_dir: Path, sample_type: str):
    plt.rcParams["font.size"] = 24
    print("\n[Volcano] Generating bubble plots (design 1)", flush = True)
    plt.close('all')
    csv_input = input_dir / "CSV_files"
    if not csv_input.exists():
        print(f"[Volcano] Skipping plots: missing folder {csv_input}", flush=True)
        return
    files = [f for f in os.listdir(csv_input) if f.endswith("_FDR.csv")]

    ordered_classes = _resolve_class_order(sample_type)

    _ensure_dir(output_dir)

    for file in files:
        df = pd.read_csv(csv_input / file)

        # --- Robust checks and normalization ---
        if "UniqueID" not in df.columns or "Significance" not in df.columns:
            print(f"[Volcano] Skipping {file} (missing columns)", flush = True)
            continue

        # Normalize 'Significance' labels (avoid trailing spaces or lowercase)
        df["Significance"] = df["Significance"].astype(str).str.strip().str.capitalize()

        # Ensure expected values exist
        valid_sig = df["Significance"].isin(["Up", "Down"])
        print(f"[Volcano] {file}: {valid_sig.sum()} significant entries detected", flush = True)
        if valid_sig.sum() == 0:
            print(f"[Volcano] {file}: 0 significant entries; producing empty plot", flush=True)
            # Create an empty but labeled figure so folders are consistent
            plt.text(0.5, 0.5, "No significant features\n(at current thresholds)",
                    ha="center", va="center")
            plt.axis("off")
            outname = output_dir / file.replace("_FDR.csv", "_EMPTY.png")
            _save_current_figure(outname, dpi=120)
            plt.close()
            # and continue to next file
            continue

        if "Lipid Class" in df.columns:
            df["Class_Group"] = (
                df["Lipid Class"]
                .astype(str)
                .str.strip()
                .replace("", "Unknown")
                .replace("nan", "Unknown")
            )
        else:
            df["Class"] = df["UniqueID"].apply(_extract_Class)
            df["Class_Group"] = df["Class"].map(_CLASS_GROUP_MAP).fillna(df["Class"])

        sig_df = df[df["Significance"].isin(["Up", "Down"])].copy()
        sig_df = sig_df[
            sig_df["log2(Fold Change)"].notna() &
            sig_df["FDR p-value"].notna() &
            sig_df["Class_Group"].notna()
        ]
        sig_df = sig_df[sig_df["Class_Group"].apply(lambda x: isinstance(x, str))]
        if sig_df.empty:
            print(f"[Volcano] {file}: no significant lipids to plot.", flush = True)
            continue
        sig_df["Class_Group"] = sig_df["Class_Group"].map(_canon_class)

        # Map Classes to indices for y-axis
        y_mapping = {cls: i for i, cls in enumerate(ordered_classes)}
        sig_df["y"] = sig_df["Class_Group"].map(y_mapping).astype(float)

        sig_df["Class_Group"] = sig_df["Class_Group"].map(_canon_class)
        class_palette = _build_class_palette(sig_df["Class_Group"].astype(str).tolist(), sample_type=sample_type)
        fc_vals = sig_df["log2(Fold Change)"].to_numpy()
        if np.isfinite(fc_vals).sum() == 0:
            continue

        lo, hi = np.nanpercentile(fc_vals, [1, 99])
        if not np.isfinite(lo) or not np.isfinite(hi):
            lo, hi = -1.0, 1.0

        span = max(0.5, hi - lo)
        vmin = lo - 0.05 * span
        vmax = hi + 0.05 * span

        # --- Bubble sizes from FDR with percentile caps ---
        fdr_log10 = -np.log10(sig_df["FDR p-value"].clip(lower=1e-300, upper=1.0)).to_numpy()
        s_lo, s_hi = np.nanpercentile(fdr_log10, [5, 95])
        if not np.isfinite(s_lo): s_lo = 0.0
        if not np.isfinite(s_hi) or s_hi <= s_lo: s_hi = max(s_lo + 1.0, 2.0)
        fdr_capped = np.clip(fdr_log10, s_lo, s_hi)
        # Map to a consistent on-screen size range
        bubble_sizes = np.interp(fdr_capped, [s_lo, s_hi], [30, 600])

        fig, ax = plt.subplots(figsize=(10, max(12, 0.4 * len(ordered_classes))))
        ax.scatter(
            sig_df["log2(Fold Change)"].to_numpy(), sig_df["y"].to_numpy(),
            s=bubble_sizes,
            c=[class_palette.get(cls, (0.5, 0.5, 0.5)) for cls in sig_df["Class_Group"].astype(str)],
            edgecolor="white", linewidth=0.4, alpha=0.85
        )

        ax.set_yticks(list(y_mapping.values()))
        ax.set_yticklabels(list(y_mapping.keys()), fontsize = 18)
        ax.set_ylim(-0.5, len(ordered_classes) - 0.5)

        # Robust x-limits
        ax.set_xlim(vmin, vmax)
        ax.axvline(0, color="gray", linestyle="--", linewidth=0.6)
        ax.tick_params(axis='x', labelsize=18)
        
        # Alternating horizontal background bands for class readability
        for i in range(len(ordered_classes)):
            if i % 2 == 0:
                ax.axhspan(i - 0.5, i + 0.5, color="lightgray", alpha=0.10, zorder=0)
        ax.set_axisbelow(True)

        ax.set_xlabel("log2(Fold Change)", fontsize=22, labelpad=10)
        ax.set_title(f"Fold-change distribution for significantly altered lipids\n({file.replace('_FDR.csv', '')})", pad=20, fontsize=22)

        # Bubble size legend (two reference sizes based on the mapping)
        legend_sizes = [np.interp(v, [s_lo, s_hi], [30, 600]) for v in [s_lo, s_hi]]
        legend_labels = [f"FDR≈{10**(-s_lo):.2g}", f"FDR≈{10**(-s_hi):.2g}"]
        size_handles = [
            Line2D([0], [0], marker="o", color="none", label=legend_labels[i],
                   markerfacecolor="gray", markeredgecolor="white", markersize=np.sqrt(s/np.pi))
            for i, s in enumerate(legend_sizes)
        ]
        class_handles = [
            Patch(facecolor=class_palette[cls], edgecolor="black", linewidth=0.3, label=cls)
            for cls in class_palette
            if cls in sig_df["Class_Group"].astype(str).tolist()
        ]
        # class_legend = ax.legend(handles=class_handles, fontsize=16,
        #                          bbox_to_anchor=(1.2, 1), loc="upper left", borderaxespad=0.0, frameon=False)
        # ax.add_artist(class_legend)
        ax.legend(handles=size_handles, title="FDR (size)\n", fontsize=14, title_fontsize=14, 
                  bbox_to_anchor=(1.28, 0.38),
                  loc="upper left", borderaxespad=0.0, frameon=False)

        plt.tight_layout()

        # Add full rectangular border
        for spine in ax.spines.values():
            spine.set_visible(True)
            spine.set_linewidth(1.0)
            spine.set_color("black")
            
        outname = output_dir / file.replace("_FDR.csv", "_bubble_plot.png")
        _save_figure(fig, outname, dpi=120, bbox_inches="tight", facecolor=fig.get_facecolor())
        plt.close(fig)


# ==========================================================
# Bubble plot (design 2): x = Class (jitter), color by Class palette
# ==========================================================
def _generate_bubble_plot_from_csv_design2(input_dir: Path, output_dir: Path, sample_type: str):
    plt.rcParams["font.size"] = 24
    print("\n[Volcano] Generating bubble plots (design 2)", flush = True)
    plt.close('all')
    csv_input = input_dir / "CSV_files"
    if not csv_input.exists():
        print(f"[Volcano] Skipping plots: missing folder {csv_input}", flush=True)
        return
    files = [f for f in os.listdir(csv_input) if f.endswith("_FDR.csv")]

    ordered_classes = _resolve_class_order(sample_type)

    _ensure_dir(output_dir)

    for file in files:
        df = pd.read_csv(csv_input / file)

        # --- Robust checks and normalization ---
        if "UniqueID" not in df.columns or "Significance" not in df.columns:
            print(f"[Volcano] Skipping {file} (missing columns)", flush = True)
            continue

        # Normalize 'Significance' labels (avoid trailing spaces or lowercase)
        df["Significance"] = df["Significance"].astype(str).str.strip().str.capitalize()

        # Ensure expected values exist
        valid_sig = df["Significance"].isin(["Up", "Down"])
        print(f"[Volcano] {file}: {valid_sig.sum()} significant entries detected", flush = True)
        if valid_sig.sum() == 0:
            print(f"[Volcano] {file}: 0 significant entries; producing empty plot", flush=True)
            # Create an empty but labeled figure so folders are consistent
            plt.text(0.5, 0.5, "No significant features\n(at current thresholds)",
                    ha="center", va="center")
            plt.axis("off")
            outname = output_dir / file.replace("_FDR.csv", "_EMPTY.png")
            _save_current_figure(outname, dpi=120)
            plt.close()
            # and continue to next file
            continue

        if "Lipid Class" in df.columns:
            df["Class_Group"] = (
                df["Lipid Class"]
                .astype(str)
                .str.strip()
                .replace("", "Unknown")
                .replace("nan", "Unknown")
            )
        else:
            df["Class"] = df["UniqueID"].apply(_extract_Class)
            df["Class_Group"] = df["Class"].map(_CLASS_GROUP_MAP).fillna(df["Class"])
        
        sig_df = df[df["Significance"].isin(["Up", "Down"])].copy()
        sig_df = sig_df[
            sig_df["log2(Fold Change)"].notna() &
            sig_df["FDR p-value"].notna() &
            sig_df["Class_Group"].notna()
        ]
        sig_df = sig_df[sig_df["Class_Group"].apply(lambda x: isinstance(x, str))]
        if sig_df.empty:
            print(f"[Volcano] {file}: no significant lipids to plot.", flush = True)
            continue

               # X positions (categorical → numeric + jitter)
        x_mapping = {cls: i for i, cls in enumerate(ordered_classes)}
        sig_df["x"] = sig_df["Class_Group"].map(x_mapping).astype(float)
        # Keep jitter tight so points don't get clipped at figure edges
        sig_df["x_jittered"] = sig_df["x"] + np.random.uniform(-0.22, 0.22, size=len(sig_df))

        # Robust bubble sizes from FDR (percentile caps)
        fdr_log10 = -np.log10(sig_df["FDR p-value"].clip(lower=1e-300, upper=1.0)).to_numpy()
        s_lo, s_hi = np.nanpercentile(fdr_log10, [5, 95])
        if not np.isfinite(s_lo): s_lo = 0.0
        if not np.isfinite(s_hi) or s_hi <= s_lo: s_hi = max(s_lo + 1.0, 2.0)
        fdr_capped = np.clip(fdr_log10, s_lo, s_hi)
        bubble_sizes = np.interp(fdr_capped, [s_lo, s_hi], [30, 600])

        class_palette = _build_class_palette(sig_df["Class_Group"].astype(str).tolist(), sample_type=sample_type)
        colors = [class_palette.get(cls, (0.5, 0.5, 0.5)) for cls in sig_df["Class_Group"]]

        fig, ax = plt.subplots(figsize=(15, max(6, 0.3 * len(ordered_classes))))
        ax.scatter(
            sig_df["x_jittered"].to_numpy(), sig_df["log2(Fold Change)"].to_numpy(),
            s=bubble_sizes, c=colors, edgecolor="white", linewidth=0.4, alpha=0.7
        )

        # Alternating vertical background bands for class readability
        for i, xpos in enumerate(x_mapping.values()):
            if i % 2 == 0:
                ax.axvspan(xpos - 0.5, xpos + 0.5, color="lightgray", alpha=0.10, zorder=0)
        ax.set_axisbelow(True)
        
        # Robust symmetric y-limits from percentiles of |log2FC|
        fc_vals = sig_df["log2(Fold Change)"].to_numpy()
        lo, hi = np.nanpercentile(np.abs(fc_vals), [1, 99])
        ymax = max(0.5, hi)
        ax.set_ylim(-ymax * 1.1, ymax * 1.1)
        ax.axhline(0, color="gray", linestyle="--", linewidth=0.6)

        ax.spines["top"].set_visible(False)
        ax.spines["right"].set_visible(False)
        ax.spines["left"].set_visible(True)
               
        ax.set_ylabel("log2(Fold Change)", fontsize=22, labelpad=10)
        ax.set_title(
            f"Fold-change distribution for significantly altered lipids\n({file.replace('_FDR.csv', '')})",
            fontsize=22,
            pad=20,)

        ax.set_xticks(list(x_mapping.values()))
        ax.set_xticklabels(list(x_mapping.keys()), rotation=90, fontsize = 16)
        ax.set_xlim(-0.5, len(ordered_classes) - 0.5)
        ax.tick_params(axis='y', labelsize=18)
        

        # Size legend consistent with mapping
        legend_sizes = [np.interp(v, [s_lo, s_hi], [30, 600]) for v in [s_lo, s_hi]]
        legend_labels = [f"FDR≈{10**(-s_lo):.2g}", f"FDR≈{10**(-s_hi):.2g}"]
        size_handles = [
            Line2D([0], [0], marker="o", color="none", label=legend_labels[i],
                   markerfacecolor="gray", markeredgecolor="white", markersize=np.sqrt(s/np.pi))
            for i, s in enumerate(legend_sizes)
        ]
        class_handles = [
            Patch(facecolor=class_palette[cls], edgecolor="black", linewidth=0.3, label=cls)
            for cls in class_palette
        ]
        class_legend = ax.legend(handles=class_handles, fontsize=16,
                                 bbox_to_anchor=(1.05, 1), loc="upper left", frameon=False)
        # ax.add_artist(class_legend)
        ax.legend(handles=size_handles, title="FDR (size)\n", fontsize=14, title_fontsize=14,
                  bbox_to_anchor=(1.03, 0.4), loc="upper left", frameon=False)

        plt.tight_layout()

        # Add full rectangular border
        for spine in ax.spines.values():
            spine.set_visible(True)
            spine.set_linewidth(1.0)
            spine.set_color("black")
            
        outname = output_dir / file.replace("_FDR.csv", "_bubble_plot_v2.png")
        _save_figure(fig, outname, dpi=120, bbox_inches="tight", facecolor=fig.get_facecolor())
        plt.close(fig)


# ==========================================================
# Public entry point (called by GUI)
# ==========================================================
def run_volcano(file_path, group_file, save_dir,
                method="fdr_bh", test_type="welch",
                run_bar_plots=True, run_bubble_plots=True,
                sample_type="Mammalians",
                p_value_threshold=0.05, fdr_threshold=0.05, fold_change_threshold=1.5, 
                group_colors=None, group_order=None,
                selected_comparisons: Optional[list[tuple[str, str]]] = None,
                annotate_labels: bool = False,
                dpi: int = 100, publication_theme: bool = False):
    """
    Main entry point for Volcano + summaries + optional bar/bubble plots.

    Args mirror your previous script; outputs and folder structure remain the same.
    """
    file_path = Path(file_path)
    save_dir = prepare_output_dir(Path(save_dir))
    volcano_dir = _ensure_dir(save_dir)
    csv_dir = _ensure_dir(volcano_dir / "CSV_files")
    plt.close('all')
    style = get_figure_style(publication_theme=publication_theme, dpi=dpi)

    print(f"[Volcano] Starting for: {file_path.name}", flush = True)
    print(f"[Volcano] sample_type received: {sample_type}", flush=True)

    # Load standardized dataset
    # X: samples × features (columns = UniqueID or UniqueIDs)
    # y: sample groups
    # Load data
    X, y, feature_meta = load_dataset(file_path, group_file)
    
    # Ensure alignment and string dtype for group ops
    X = X.reset_index(drop=True)
    y = y.astype(str).str.strip().reset_index(drop=True)

    # -------------------------------------------------------
    # Build metadata lookup table from feature_meta
    # -------------------------------------------------------
    meta_lookup = None
    if isinstance(feature_meta, pd.DataFrame) and not feature_meta.empty:
        meta_lookup = feature_meta.copy()

        # Canonicalize column names
        rename_cols = {}
        for c in meta_lookup.columns:
            cl = str(c).strip().lower()
            if cl in {"uniqueid", "unique id", "feature id", "id"}:
                rename_cols[c] = "UniqueID"
            elif cl in {"annotation"}:
                rename_cols[c] = "Annotation"
            elif cl in {"annotation type", "annotation_type"}:
                rename_cols[c] = "Annotation Type"
            elif cl in {"headgroup", "head group"}:
                rename_cols[c] = "Headgroup"
            elif cl in {"lipid class", "lipid_class", "class"}:
                rename_cols[c] = "Lipid Class"
            elif cl in {"number of carbons in fatty acyls", "total carbons", "carbons", "ncarbons", "n_carbons"}:
                rename_cols[c] = "Number of carbons in fatty acyls"
            elif cl in {"double bond equivalents", "double bonds", "db", "dbe"}:
                rename_cols[c] = "Double bond equivalents"
        if rename_cols:
            meta_lookup = meta_lookup.rename(columns=rename_cols)

        # Split combined column if present
        combo = [c for c in meta_lookup.columns
                if str(c).strip().lower() in {"annotation type headgroup", "annotation type and headgroup"}]
        if combo and ("Annotation Type" not in meta_lookup.columns or "Headgroup" not in meta_lookup.columns):
            col = combo[0]
            s = meta_lookup[col].astype(str).fillna("")
            parts = s.str.split(r"\s*[|;:]\s*", n=1, expand=True)
            if parts.shape[1] == 1:
                parts = s.str.split(r"\s{2,}", n=1, expand=True)
            meta_lookup["Annotation Type"] = parts[0].astype(str).str.strip()
            meta_lookup["Headgroup"] = (parts[1].astype(str).str.strip() if parts.shape[1] > 1 else "")

        # Ensure required columns exist
        if "UniqueID" not in meta_lookup.columns:
            raise ValueError("[Volcano] feature_meta lacks a UniqueID column (or known alias).")

        for c in ["Annotation", "Annotation Type", "Headgroup", "Lipid Class", "Number of carbons in fatty acyls", "Double bond equivalents"]:
            if c not in meta_lookup.columns:
                meta_lookup[c] = np.nan if c in {"Number of carbons in fatty acyls", "Double bond equivalents"} else ""

        # Clean
        meta_lookup["UniqueID"] = meta_lookup["UniqueID"].astype(str).str.strip()
        for c in ["Annotation", "Annotation Type", "Headgroup", "Lipid Class"]:
            meta_lookup[c] = (
                meta_lookup[c]
                .astype(str)
                .replace({"nan": "", "NaN": "", "None": "", "<NA>": "", "NA": ""})
                .str.strip()
            )
        for c in ["Number of carbons in fatty acyls", "Double bond equivalents"]:
            meta_lookup[c] = pd.to_numeric(meta_lookup[c], errors="coerce")

        meta_lookup = meta_lookup[[
            "UniqueID", "Annotation", "Annotation Type", "Headgroup", "Lipid Class",
            "Number of carbons in fatty acyls", "Double bond equivalents"
        ]].drop_duplicates("UniqueID")
    else:
        meta_lookup = pd.DataFrame(columns=[
            "UniqueID", "Annotation", "Annotation Type", "Headgroup", "Lipid Class",
            "Number of carbons in fatty acyls", "Double bond equivalents"
        ])
    # -------------------------------------------------------

    # --- Skip QC samples completely --------------------------------------
    mask_non_qc = ~y.str.lower().str.contains("qc", na=False)
    X = X.loc[mask_non_qc].reset_index(drop=True)
    y = y.loc[mask_non_qc].reset_index(drop=True)
    # ---------------------------------------------------------------------

    # Drop empty / nan-like group labels
    bad = y.str.lower().isin(["", "nan", "none", "<na>", "na", "null"])
    X = X.loc[~bad].reset_index(drop=True)
    y = y.loc[~bad].reset_index(drop=True)

    groups = sorted(y.unique().tolist())
    print(f"[Volcano] Groups (excluding QC): {groups}", flush=True)

    all_comparisons = [(g1, g2) for g1 in groups for g2 in groups if str(g1) != str(g2)]
    selected_pairs = []
    if selected_comparisons:
        available_set = {(str(g1), str(g2)) for g1, g2 in all_comparisons}
        seen = set()
        for g1, g2 in selected_comparisons:
            pair = (str(g1).strip(), str(g2).strip())
            if not pair[0] or not pair[1] or pair[0] == pair[1]:
                continue
            if pair not in available_set or pair in seen:
                continue
            selected_pairs.append(pair)
            seen.add(pair)
        if not selected_pairs:
            raise ValueError(
                f"[Volcano] None of the selected comparisons matched the available groups: {groups}"
            )
        print(f"[Volcano] Running selected comparisons only: {selected_pairs}", flush=True)
    else:
        selected_pairs = all_comparisons
        print(f"[Volcano] Running all pairwise comparisons ({len(selected_pairs)})", flush=True)

    # Pairwise comparisons (directional: A vs B and B vs A)
    for g1, g2 in selected_pairs:

            print(f"[Volcano] Comparing {g1} vs {g2}", flush=True)
            df_volc = _compute_volcano(
                g1, g2, X, y, meta_lookup,
                method=method, test_type=test_type,
                p_thresh=p_value_threshold, fdr_thresh=fdr_threshold,
                fc_thresh=fold_change_threshold,
                assumption_alpha=0.05
            )

            # Sanitize before saving/plotting
            df_volc = df_volc.replace([np.inf, -np.inf], np.nan)
            df_volc = df_volc.dropna(subset=["Fold Change", "log2(Fold Change)", "p-value", "FDR p-value"])

            # --- log counts clearly ---
            up_n = int((df_volc["Significance"] == "Up").sum())
            down_n = int((df_volc["Significance"] == "Down").sum())
            print(f"[Volcano] {g1}_vs_{g2}: {up_n} Up, {down_n} Down (FC≥{fold_change_threshold:.2f}, "
                  f"p<{p_value_threshold}, FDR<{fdr_threshold})", flush=True)

            csv_out = csv_dir / f"{_sanitize_filename(g1)}_vs_{_sanitize_filename(g2)}_FDR.csv"
            all_tests_csv_out = csv_dir / f"{_sanitize_filename(g1)}_vs_{_sanitize_filename(g2)}_all_test_pvalues.csv"

            # Always write a file so summaries see every comparison
            if df_volc.empty:
                empty_cols = [
                    "UniqueID", "Annotation", "Annotation Type", "Headgroup", "Lipid Class",
                    "Fold Change", "log2(Fold Change)", "p-value", "FDR p-value",
                    "-log10(FDR p-value)", "Significance",
                ]
                pd.DataFrame(columns=empty_cols).to_csv(csv_out, index=False)
                pd.DataFrame(columns=[
                    "UniqueID", "Annotation", "Annotation Type", "Headgroup", "Lipid Class",
                    "Number of carbons in fatty acyls", "Double bond equivalents",
                    "n_group1", "n_group2",
                    "Auto selected test",
                    "p-value (Auto)", "p-value (Student)", "p-value (Welch)", "p-value (Mann-Whitney)",
                    "Mann-Whitney rank-biserial", "Mann-Whitney direction agrees", "Mann-Whitney robustness",
                    "p-value",
                ]).to_csv(all_tests_csv_out, index=False, encoding="utf-8-sig")
            else:
                df_volc.to_csv(csv_out, index=False)
                _export_all_test_pvalues(df_volc, all_tests_csv_out)

            # Always plot (function will render an empty placeholder if needed)
            _plot_volcano(
                df_volc, g1, g2, volcano_dir,
                p_thresh=p_value_threshold,
                fdr_thresh=fdr_threshold,
                fc_thresh=fold_change_threshold,
                style=style,
                sample_type=sample_type,
                annotate_labels=annotate_labels,
            )

    # Summary tables
    _save_summary_tables(
        csv_dir=csv_dir,
        output_csv=volcano_dir / "Summary_table_raw.csv",
        output_excel=volcano_dir / "Summary_Volcano.xlsx",
        fc_thresh=fold_change_threshold, fdr_thresh=fdr_threshold, p_thresh=p_value_threshold
    )
    _save_class_enrichment_summary(
        csv_dir=csv_dir,
        output_csv=volcano_dir / "volcano_class_enrichment.csv",
    )

    # class bar plots
    if run_bar_plots:
        bar_dir = _ensure_dir(volcano_dir / "Bar_plots")
        _generate_class_barplots_from_csv(
            input_dir=volcano_dir, output_dir=bar_dir, sample_type=sample_type,
            p_value_threshold=p_value_threshold, fdr_threshold=fdr_threshold, fold_change_threshold=fold_change_threshold
        )

    # bubble plots (both designs)
    if run_bubble_plots:
        bubble_dir = _ensure_dir(volcano_dir / "Bubble_plots")
        _generate_bubble_plot_from_csv(
            input_dir=volcano_dir, output_dir=bubble_dir, sample_type=sample_type
        )
        _generate_bubble_plot_from_csv_design2(
            input_dir=volcano_dir, output_dir=bubble_dir, sample_type=sample_type
        )

    complementary_dir = _ensure_dir(volcano_dir / "Complementary_plots")
    _generate_ranked_effect_plots_from_csv(
        input_dir=volcano_dir, output_dir=complementary_dir, sample_type=sample_type
    )
    _generate_class_directionality_plots_from_csv(
        input_dir=volcano_dir, output_dir=complementary_dir, sample_type=sample_type
    )

    print(f"[Volcano] Completed. Output in: {volcano_dir}\n", flush = True)


# Optional local test
if __name__ == "__main__":
    import sys
    fpath = Path(sys.argv[1]) if len(sys.argv) > 1 else Path.cwd() / "statistics" / "Final_Annotated.csv"
    gpath = Path(sys.argv[2]) if len(sys.argv) > 2 else Path.cwd() / "statistics" / "sample_groups_cleaned.csv"
    outdir = Path(sys.argv[3]) if len(sys.argv) > 3 else Path.cwd() / "statistics" / "Volcano" / "ManualTest"

    run_volcano(
        file_path=fpath,
        group_file=gpath if gpath.exists() else None,
        save_dir=outdir,
        method="fdr_bh",
        test_type="welch",
        run_bar_plots=True,
        run_bubble_plots=True,
        sample_type="Mammalians",
        p_value_threshold=0.05,
        fdr_threshold=0.10,
        fold_change_threshold=1.5,
    )
